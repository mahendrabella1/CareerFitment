#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Builds the class 11-12 stream / degree / eligibility data from
"11-12th Streams list.xlsx".

The workbook has two kinds of tab:

  * "Class 12 Streams" - four stream FAMILIES (Science, Commerce,
    Humanities/Arts, Vocational) laid out in columns, each listing the subject
    combinations a student can actually be enrolled in (PCM, PCB, Commerce with
    Mathematics, and so on). This is what the student picks first.

  * Eight domain tabs (Engineering, Medical, Pure Sciences, Commerce, Law,
    Humanities, Design, Other) - a matrix of DEGREE x STREAM whose cells hold a
    traffic-light marker:

        green   directly eligible
        yellow  possible, but with a condition (bridge subject, entrance
                exam, or institution-specific rule)
        red     not eligible from that stream

    Cells can hold a PAIR ("red/yellow") meaning it depends on the institution.
    We take the BEST of the pair and record that it is conditional, because
    telling a student "not possible" when some universities do allow it is the
    more damaging error of the two.

A degree column header names a stream GROUP (PCM, PCB, Commerce, ...), not the
specific combination a student picked. Combination -> group is resolved by
STREAM_GROUP below, so "PCM + Economics" inherits the PCM column.

The "Other" tab is a plain two-column list of fields with no eligibility matrix
at all, so those degrees are recorded as open to every stream - which matches
reality for media and hospitality programmes.

Writes:  data/demo-11-12/streams.json

Run:     python scripts/build_demo_streams.py
"""

import json
import os
import sys

HERE = os.path.dirname(os.path.abspath(__file__))
PROJECT = os.path.join(HERE, "..")
DATA = os.path.join(PROJECT, "data", "demo-11-12")
XLSX = os.path.join(PROJECT, "..", "11-12th Streams list.xlsx")

sys.path.insert(0, HERE)
from demo_11_12_extra_streams import as_combinations, family_additions  # noqa: E402

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required:  pip install openpyxl")

GREEN, YELLOW, RED = "\U0001f7e2", "\U0001f7e1", "\U0001f534"
RANK = {"green": 3, "yellow": 2, "red": 1}


def read_marker(raw):
    """A cell to (verdict, conditional). Pairs collapse to their better half."""
    s = "" if raw is None else str(raw).strip()
    if not s:
        return None, False
    found = []
    if GREEN in s:
        found.append("green")
    if YELLOW in s:
        found.append("yellow")
    if RED in s:
        found.append("red")
    if not found:
        # Some cells carry prose instead of a marker, e.g. "Later pathway" on
        # the LLB-after-graduation row. Those are genuinely open to everyone,
        # just not straight after class 12.
        return ("green", True) if s else (None, False)
    best = max(found, key=lambda f: RANK[f])
    return best, len(found) > 1 or best == "yellow"


# Column headers in the domain tabs name a stream GROUP. A student picks a
# specific combination; this maps one to the other.
STREAM_GROUP = {
    # Science
    "PCM / MPC": "PCM",
    "PCB / BiPC": "PCB",
    "PCMB / MBiPC": "PCMB",
    "PCMC / PCM + Computer Science": "PCM",
    "PCM + Electronics": "PCM",
    "PCM + Economics": "PCM",
    "PCB + Psychology": "PCB",
    "PCB + Biotechnology": "PCB",
    "PCB + Computer Science": "PCB",
    "PCMB + Computer Science": "PCMB",
    # Commerce
    "Commerce without Mathematics": "Commerce no Math",
    "Commerce with Mathematics": "Commerce+Math",
    "Commerce + Computer Science": "Commerce no Math",
    "Commerce + Economics + Mathematics": "Commerce+Math",
    "Commerce + Applied Mathematics": "Commerce+Math",
    # Humanities
    "Humanities without Mathematics": "Humanities no Math",
    "Humanities with Mathematics": "Humanities+Math",
    "Humanities + Economics": "Humanities no Math",
    "Humanities + Psychology": "Humanities no Math",
    "Humanities + Computer Science": "Humanities no Math",
    "Humanities + Legal Studies": "Humanities no Math",
    "Humanities + Fine Arts": "Humanities no Math",
}

# Domain-tab column headers vary ("Commerce" on one tab, "Commerce+Math" on
# another). A group falls back through this chain until the tab has a column.
GROUP_FALLBACK = {
    "PCM": ["PCM"],
    "PCB": ["PCB"],
    "PCMB": ["PCMB", "PCM"],
    "Commerce+Math": ["Commerce+Math", "Commerce"],
    "Commerce no Math": ["Commerce no Math", "Commerce"],
    "Humanities+Math": ["Humanities+Math", "Humanities"],
    "Humanities no Math": ["Humanities no Math", "Humanities"],
    "Vocational": ["Vocational"],
}

FAMILY_OF = {
    "PCM": "Science", "PCB": "Science", "PCMB": "Science",
    "Commerce+Math": "Commerce", "Commerce no Math": "Commerce",
    "Humanities+Math": "Humanities", "Humanities no Math": "Humanities",
    "Vocational": "Vocational",
}


def parse_stream_families(ws):
    """The four-column layout of the "Class 12 Streams" tab."""
    rows = [[("" if c is None else str(c).strip()) for c in r] for r in ws.iter_rows(values_only=True)]
    header = rows[0]
    families = []
    for col, name in enumerate(header):
        if not name:
            continue
        combos = []
        for r in rows[1:]:
            if col < len(r) and r[col]:
                combos.append(r[col])
        families.append({"family": name, "combinations": combos})
    return families


def parse_domain_tab(ws, domain):
    """Domain tabs hold one or more matrix BLOCKS, each with its own header row
    naming the stream columns. Blocks are separated by a title row and blanks."""
    rows = [[("" if c is None else str(c).strip()) for c in r] for r in ws.iter_rows(values_only=True)]
    degrees = []
    header = None
    for r in rows:
        if not any(r):
            header = None
            continue
        first = r[0]
        # A header row is the one whose first cell names the row type.
        if first in ("Degree family", "Degree", "Programme", "Qualification"):
            header = r
            continue
        if header is None:
            continue  # a block title such as "Architecture & Planning"
        if not first:
            continue
        elig = {}
        for i, col in enumerate(header):
            if i == 0 or not col:
                continue
            verdict, conditional = read_marker(r[i] if i < len(r) else "")
            if verdict:
                elig[col] = {"verdict": verdict, "conditional": conditional}
        if elig:
            degrees.append({"degree": first.rstrip("*").strip(), "domain": domain, "eligibility": elig})
    return degrees


def parse_other_tab(ws):
    """Two independent lists side by side, no eligibility matrix."""
    rows = [[("" if c is None else str(c).strip()) for c in r] for r in ws.iter_rows(values_only=True)]
    out = []
    header = rows[0]
    for col, domain in enumerate(header):
        if not domain:
            continue
        for r in rows[1:]:
            if col < len(r) and r[col]:
                out.append({"degree": r[col], "domain": domain, "eligibility": {}})
    return out


# Which family a matrix COLUMN belongs to. Used to tell "the workbook says no"
# apart from "the workbook never considered this stream".
COLUMN_FAMILY = {
    "PCM": "Science", "PCB": "Science", "PCMB": "Science",
    "Commerce": "Commerce", "Commerce+Math": "Commerce", "Commerce no Math": "Commerce",
    "Humanities": "Humanities", "Humanities+Math": "Humanities",
    "Humanities no Math": "Humanities",
    "Vocational": "Vocational",
}


def resolve(degree, group):
    """Verdict for one degree from one stream group.

    Three distinct cases, and conflating them is how MBBS ends up looking open
    to a commerce student:

    1. No matrix at all (the Other tab: media, hospitality). Those lists carry
       no eligibility columns because the programmes really are open to every
       stream. -> eligible.

    2. Matrix has a column for this stream (directly or via GROUP_FALLBACK).
       -> whatever the workbook says.

    3. Matrix has columns, but none for this student FAMILY. The domain tab
       enumerated the streams it accepts and this was not among them: the
       Medical tab lists only PCM/PCB/PCMB, which is precisely how it says
       "medicine is a science-stream pathway". -> not eligible, flagged
       `unlisted` so the UI can explain it as "not offered from your stream"
       rather than implying the student failed some requirement.
    """
    elig = degree["eligibility"]
    if not elig:
        return "green", True, False
    for col in GROUP_FALLBACK.get(group, [group]):
        if col in elig:
            e = elig[col]
            return e["verdict"], e["conditional"], False
    family = FAMILY_OF.get(group)
    covered = {COLUMN_FAMILY.get(c) for c in elig}
    if family in covered:
        # The family is on the tab but this exact combination is not spelled
        # out. Genuinely unknown, so say "check the institution" rather than
        # inventing either answer.
        return "yellow", True, False
    return "red", False, True


def main():
    if not os.path.exists(XLSX):
        sys.exit("workbook not found: " + XLSX)
    wb = openpyxl.load_workbook(XLSX, data_only=True)

    families = parse_stream_families(wb["Class 12 Streams"])
    domain_tabs = {
        "Eng. and Tech": "Engineering & Technology",
        "Medical & Health Sciences": "Medical & Health Sciences",
        "Pure Sciences": "Pure Sciences & Agriculture",
        "Commerce": "Commerce & Management",
        "Law": "Law",
        "Humanities or Social Sciences": "Humanities & Social Sciences",
        "Design": "Design",
    }
    degrees = []
    for tab, domain in domain_tabs.items():
        degrees.extend(parse_domain_tab(wb[tab], domain))
    degrees.extend(parse_other_tab(wb["Other"]))

    # De-duplicate: a few degrees appear on two tabs (B.Des on Engineering and
    # on Design, B.Tech Biotechnology on Engineering and Pure Sciences). Keep
    # the first and merge the eligibility columns of the rest, so no column is
    # lost just because the degree was listed twice.
    merged = {}
    for d in degrees:
        key = d["degree"]
        if key in merged:
            for col, v in d["eligibility"].items():
                merged[key]["eligibility"].setdefault(col, v)
        else:
            merged[key] = d
    degrees = list(merged.values())

    # Combination -> stream GROUP only. The per-degree verdicts used to be
    # expanded here: 32 combinations x 118 degrees, each row repeating the
    # degree name and domain. That was 478KB of a 513KB file, all of it
    # derivable, and it pushed the Next build past its default heap.
    #
    # It was also a second copy of the same truth. Eligibility is now resolved
    # once at request time by lib/demo11/eligibility.ts from `degrees` and the
    # lookup tables emitted below, so there is exactly one place it can be
    # wrong.
    combos = []
    for fam in families:
        for combo in fam["combinations"]:
            group = STREAM_GROUP.get(combo, "Vocational" if fam["family"] == "Vocational" else combo)
            combos.append({"combination": combo, "family": fam["family"], "group": group})

    # The AP/TS intermediate combinations the workbook omits - CEC, HEC, MEC.
    # See demo_11_12_extra_streams.py. They reuse existing eligibility groups,
    # so no degree matrix changes.
    combos.extend(as_combinations())
    additions = family_additions()
    known_families = {f["family"] for f in families}
    unknown = set(additions) - known_families
    if unknown:
        sys.exit("authored stream family not on the sheet: {} (sheet has {})".format(
            ", ".join(sorted(unknown)), ", ".join(sorted(known_families))))
    for fam in families:
        for name in additions.get(fam["family"], []):
            if name not in fam["combinations"]:
                fam["combinations"].append(name)

    out = {
        "families": families,
        "groups": sorted({c["group"] for c in combos}),
        "degrees": degrees,
        "combinations": combos,
        # The tables the runtime resolver needs. Emitted as data rather than
        # duplicated in TypeScript so the two cannot drift apart.
        "groupFallback": GROUP_FALLBACK,
        "columnFamily": COLUMN_FAMILY,
        "familyOf": FAMILY_OF,
    }
    os.makedirs(DATA, exist_ok=True)
    path = os.path.join(DATA, "streams.json")
    with open(path, "w", encoding="utf-8") as fh:
        json.dump(out, fh, ensure_ascii=False, indent=1)

    print("Parsed", os.path.basename(XLSX))
    authored = {c["combination"] for c in combos if c.get("authored")}
    for fam in families:
        extra = sum(1 for c in fam["combinations"] if c in authored)
        note = "  (+{} authored)".format(extra) if extra else ""
        print("  {:16s} {:2d} combinations{}".format(
            fam["family"], len(fam["combinations"]), note))
    print("  {} degrees across {} domains".format(len(degrees), len({d["domain"] for d in degrees})))
    print()
    print("Eligible degree counts by stream combination:")
    print("  (resolved here for the report only - lib/demo11/eligibility.ts is")
    print("   the authority at runtime and implements the same rules)")
    for c in combos:
        verdicts = [resolve(d, c["group"])[0] for d in degrees]
        print("  {:38s} green {:3d}  yellow {:3d}  red {:3d}".format(
            c["combination"],
            verdicts.count("green"), verdicts.count("yellow"), verdicts.count("red")))
    print()
    print("Wrote", os.path.relpath(path, PROJECT))


if __name__ == "__main__":
    sys.stdout.reconfigure(encoding="utf-8")
    main()
