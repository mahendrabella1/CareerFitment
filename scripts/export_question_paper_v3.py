# -*- coding: utf-8 -*-
"""Export the live Class 9-10 question bank as a reviewable workbook.

Reads project/data/*.json directly — whatever a student would actually be
served — rather than any authoring source, so the sheet cannot drift from the
exam. Option ORDER matches the exam exactly, which matters now that order is
deliberately shuffled to stop the paper being answerable by position.

Writes "version 3.xlsx" beside the project folder and into project/docs.

Usage:  python scripts/export_question_paper_v3.py [output.xlsx]
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HERE = Path(__file__).resolve().parent
PROJECT = HERE.parent
DATA = PROJECT / "data"
STAGE, SET = "9-10", "Set 1"
NAME = "version 3.xlsx"

CLUSTERS = {
    "A": "Engineering & Construction", "B": "Information Technology",
    "C": "Health Science", "D": "Arts, Media & Design", "E": "Business & Marketing",
    "F": "Human & Public Services", "G": "Science, Nature & Agriculture",
    "H": "Sports, Hospitality & Lifestyle",
}
RIASEC = {"R": "Realistic", "I": "Investigative", "A": "Artistic",
          "S": "Social", "E": "Enterprising", "C": "Conventional"}

# sheet title, category, what the section measures
SECTIONS = [
    ("Interests", "career_interest", "RIASEC interest codes mapped onto 8 career clusters"),
    ("Aptitude", "aptitude", "Numerical, Verbal, Logical, Abstract, Spatial, Attention to Detail, Mechanical"),
    ("Personality", "personality", "Big Five: Openness, Conscientiousness, Extraversion, Agreeableness, Emotional Stability"),
    ("Strengths", "strengths", "Analytical, Creative, Leadership, Relationship, Execution, Communication, Adaptability, Learning"),
    ("Motivators", "motivators", "Achievement, Innovation, Impact, Leadership, Security, Learning"),
    ("Learning Styles", "learning_styles", "VARK: Visual, Aural, Read/Write, Kinesthetic"),
    ("Multiple Intelligence", "multiple_intelligence", "Gardner's eight intelligences"),
    ("Emotional Intelligence", "emotional_intelligence", "Emotional Awareness, Regulation, Empathy, Relationship Management, Adaptability"),
]

HEAD_FILL = PatternFill("solid", fgColor="1F3864")
HEAD_FONT = Font(bold=True, color="FFFFFF", size=10)
TOTAL_FILL = PatternFill("solid", fgColor="D9E2F3")
KEY_FILL = PatternFill("solid", fgColor="E2EFDA")
BODY = Font(size=10)
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP = Alignment(vertical="top", wrap_text=True)
MID = Alignment(horizontal="center", vertical="center", wrap_text=True)


def load():
    b = json.loads((DATA / "assessment-questions.json").read_text(encoding="utf-8"))
    b["aptitude"] = json.loads((DATA / "aptitude-questions.json").read_text(encoding="utf-8"))
    b["strengths"] = json.loads((DATA / "strengths-questions.json").read_text(encoding="utf-8"))
    return {c: b[c][STAGE][SET] for c in b}


# --------------------------------------------------------------------------- #
# Career clusters and professions PER OPTION, for every section.
#
# Only the interests questions name careers directly. Everywhere else an option
# scores a construct — a Big Five trait, a strength, a motivator — and the career
# map says which professions that construct predicts. So the careers shown here
# are not decoration: they are read out of the same affinity table the scorer
# uses, which means the sheet shows what each option actually contributes to a
# recommendation.
#
# Learning styles is deliberately absent from the career map (the workbook's own
# logic tab: this dimension must never change career recommendations, it only
# drives study advice), so it is reported as such rather than left blank.
# --------------------------------------------------------------------------- #
CAREER_MAP = json.loads((DATA / "career-map-9-10.json").read_text(encoding="utf-8"))
AFFINITY = CAREER_MAP["affinity"]
PROF_CLUSTER = CAREER_MAP["professionCluster"]

# category -> which affinity table its constructs live in
AFFINITY_TABLE = {
    "personality": "personality", "strengths": "strengths", "motivators": "motivators",
    "multiple_intelligence": "mi", "emotional_intelligence": "ei", "aptitude": "aptitude",
}


# Personality is the one section whose weights use letter codes; the career map
# keys on the full trait names, so the two have to be reconciled or the lookup
# silently returns nothing.
BIG_FIVE = {"O": "Openness", "C": "Conscientiousness", "E": "Extraversion",
            "A": "Agreeableness", "S": "Emotional Stability"}


def careers_for(cat: str, construct: str, limit: int = 4):
    """Professions the career map links to this construct, and their clusters."""
    table = AFFINITY_TABLE.get(cat)
    if not table:
        return [], []
    if cat == "personality":
        construct = BIG_FIVE.get(construct, construct)
    names = AFFINITY.get(table, {}).get(construct, [])
    letters = sorted({PROF_CLUSTER[n] for n in names if n in PROF_CLUSTER})
    return names[:limit], letters


def dominant(v):
    if isinstance(v, dict):
        return max(v, key=v.get) if v else "-"
    return str(v)


def option_careers(cat: str, constructs: list[str]) -> tuple[str, str]:
    """(clusters per option, professions per option) as wrapped cells."""
    cl, pr = [], []
    for c in constructs:
        names, letters = careers_for(cat, c)
        cl.append(", ".join(f"{L} {CLUSTERS[L]}" for L in letters) or "—")
        pr.append(", ".join(names) or "—")
    return per_option(cl), per_option(pr)


def points(v) -> str:
    if isinstance(v, dict):
        return ", ".join(f"{k}+{n}" for k, n in v.items()) if v else "not scored"
    return str(v)


def per_option(labels) -> str:
    return "\n".join(f"{chr(65 + i)} · {v}" for i, v in enumerate(labels))


def sheet(wb, title, headers, widths, rows, height=104):
    ws = wb.create_sheet(title)
    ws.append(headers)
    for c, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=c)
        cell.fill, cell.font, cell.alignment, cell.border = HEAD_FILL, HEAD_FONT, MID, BOX
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[1].height = 34
    for r in rows:
        ws.append(r)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
        for cell in row:
            cell.font, cell.alignment, cell.border = BODY, WRAP, BOX
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = height
        ws.cell(row=r, column=1).alignment = MID
    ws.freeze_panes = "C2"
    return ws


def build(target: Path):
    bank = load()
    wb = Workbook()
    wb.remove(wb.active)

    # ---- Blueprint
    rows, total = [], 0
    for i, (title, cat, note) in enumerate(SECTIONS, 1):
        qs = bank[cat]
        opts = sorted({len(q["options"]) for q in qs})
        first, last = qs[0]["q"], qs[-1]["q"]
        total += len(qs)
        rows.append([i, title, f"{first}-{last}", len(qs),
                     ", ".join(map(str, opts)), note])
    rows.append(["", "TOTAL", "Q1-Q60", total, "", "8 sections"])
    ws = sheet(wb, "Blueprint", ["#", "Section", "Q.No", "Questions", "Options", "Measures"],
               [5, 26, 12, 11, 10, 86], rows, height=30)
    for c in range(1, 7):
        ws.cell(row=ws.max_row, column=c).font = Font(bold=True, size=10)
        ws.cell(row=ws.max_row, column=c).fill = TOTAL_FILL

    def opt_cols(q, n=5):
        o = [str(x) for x in q["options"]]
        return o + [""] * (n - len(o))

    OPT = ["Option A", "Option B", "Option C", "Option D", "Option E"]

    # ---- Interests
    rows = []
    for q in bank["career_interest"]:
        cw, ri = q["clusterWeights"], q["riasec"]
        rows.append([q["q"], q["text"], *opt_cols(q),
                     per_option([f"{max(w, key=w.get)} — {CLUSTERS[max(w, key=w.get)]}" for w in cw]),
                     per_option([points(w) for w in cw]),
                     per_option(["+".join(sorted(r, key=r.get, reverse=True)) for r in ri]),
                     per_option([", ".join(c[:3]) for c in q.get("careers", [])])])
    sheet(wb, "Interests", ["Q.No", "Question", *OPT, "Career Cluster (per option)",
                            "Cluster Weights", "RIASEC (per option)", "Example Careers (per option)"],
          [7, 40, 30, 30, 30, 30, 30, 30, 24, 18, 44], rows, height=112)

    # ---- Aptitude
    rows = []
    for q in bank["aptitude"]:
        art = "yes" if q.get("media") else "no"
        names, letters = careers_for("aptitude", q.get("domain", ""))
        rows.append([q["q"], q["text"], *opt_cols(q), chr(65 + q["correct"]),
                     q.get("domain", ""), q.get("difficulty", ""), art,
                     ", ".join(f"{L} {CLUSTERS[L]}" for L in letters) or "-",
                     ", ".join(names) or "-"])
    ws = sheet(wb, "Aptitude", ["Q.No", "Question", *OPT, "Answer", "Dimension", "Difficulty",
                                "Has artwork", "Career Clusters", "Example Professions"],
               [7, 42, 16, 16, 16, 16, 16, 9, 20, 11, 12, 34, 40], rows, height=88)
    for r in range(2, ws.max_row + 1):
        c = ws.cell(row=r, column=8)
        c.alignment, c.font, c.fill = MID, Font(bold=True, size=10), KEY_FILL

    # ---- Personality
    rows = []
    for q in bank["personality"]:
        d = [dominant(x) for x in q["traitPoints"]]
        cl, pr = option_careers("personality", d)
        rows.append([q["q"], q["text"], *opt_cols(q), q.get("trait", ""), q.get("facet", ""),
                     per_option([BIG_FIVE.get(x, x) for x in d]),
                     per_option([points(x) for x in q["traitPoints"]]), cl, pr])
    sheet(wb, "Personality", ["Q.No", "Situation", *OPT, "Big Five Trait", "Facet",
                              "Trait (per option)", "Weights (per option)",
                              "Career Clusters (per option)", "Example Professions (per option)"],
          [7, 36, 25, 25, 25, 25, 20, 16, 16, 18, 26, 40, 46], rows, height=112)

    # ---- Strengths
    rows = []
    for q in bank["strengths"]:
        d = [dominant(x) for x in q["strengthPoints"]]
        cl, pr = option_careers("strengths", d)
        rows.append([q["q"], q["text"], *opt_cols(q), per_option(d),
                     per_option([points(x) for x in q["strengthPoints"]]), cl, pr])
    sheet(wb, "Strengths", ["Q.No", "Question", *OPT, "Strength (per option)", "Weights (per option)",
                            "Career Clusters (per option)", "Example Professions (per option)"],
          [7, 36, 26, 26, 26, 26, 26, 20, 26, 40, 46], rows, height=112)

    # ---- Motivators
    rows = []
    for q in bank["motivators"]:
        d = [dominant(x) for x in q["motivatorPoints"]]
        cl, pr = option_careers("motivators", d)
        rows.append([q["q"], q["text"], *opt_cols(q), per_option(d),
                     per_option([points(x) for x in q["motivatorPoints"]]), cl, pr])
    sheet(wb, "Motivators", ["Q.No", "Question", *OPT, "Motivator (per option)", "Weights (per option)",
                             "Career Clusters (per option)", "Example Professions (per option)"],
          [7, 36, 26, 26, 26, 26, 26, 20, 26, 40, 46], rows, height=112)

    # ---- Learning styles
    rows = [[q["q"], q["text"], *opt_cols(q), per_option(q.get("styles", [])),
             "Not used for career matching - drives study advice only"]
            for q in bank["learning_styles"]]
    sheet(wb, "Learning Styles", ["Q.No", "Scenario", *OPT, "Style (per option)", "Career Clusters"],
          [7, 42, 30, 30, 30, 30, 30, 22, 40], rows, height=100)

    # ---- Multiple intelligence
    rows = []
    for q in bank["multiple_intelligence"]:
        d = [dominant(x) for x in q["intelPoints"]]
        cl, pr = option_careers("multiple_intelligence", d)
        rows.append([q["q"], q["text"], *opt_cols(q), per_option(d),
                     per_option([points(x) for x in q["intelPoints"]]), cl, pr])
    sheet(wb, "Multiple Intelligence", ["Q.No", "Question", *OPT, "Intelligence (per option)",
                                        "Weights (per option)", "Career Clusters (per option)",
                                        "Example Professions (per option)"],
          [7, 36, 26, 26, 26, 26, 26, 24, 30, 40, 46], rows, height=112)

    # ---- Emotional intelligence
    rows = []
    for q in bank["emotional_intelligence"]:
        names, letters = careers_for("emotional_intelligence", q.get("dimension", ""))
        rows.append([q["q"], q["text"], *opt_cols(q), q.get("dimension", ""),
                     per_option([str(x) for x in q.get("scores", [])]),
                     f"{min(q['scores'])} - {max(q['scores'])}" if q.get("scores") else "",
                     ", ".join(f"{L} {CLUSTERS[L]}" for L in letters) or "-",
                     ", ".join(names) or "-"])
    sheet(wb, "Emotional Intelligence", ["Q.No", "Question", *OPT, "Dimension", "Score (per option)",
                                         "Range", "Career Clusters", "Example Professions"],
          [7, 36, 25, 25, 25, 25, 25, 24, 15, 9, 36, 42], rows, height=104)

    # ---- verify
    problems = []
    grand = 0
    for title, cat, _n in SECTIONS:
        data = [r for r in wb[title].iter_rows(min_row=2, values_only=True)
                if r[0] and str(r[0]).startswith("Q")]
        grand += len(data)
        if len(data) != len(bank[cat]):
            problems.append(f"{title}: {len(data)} rows vs {len(bank[cat])} in the bank")
        for row, q in zip(data, bank[cat]):
            if row[1] != q["text"]:
                problems.append(f"{title} {q['q']}: question text differs from the bank")
            for i, o in enumerate(q["options"]):
                if str(row[2 + i]) != str(o):
                    problems.append(f"{title} {q['q']} option {chr(65 + i)}: differs from the bank")
    if grand != 60:
        problems.append(f"exported {grand} questions, expected 60")
    if problems:
        raise SystemExit("VERIFY FAILED:\n  " + "\n  ".join(problems[:10]))

    target.parent.mkdir(parents=True, exist_ok=True)
    wb.save(target)
    print(f"wrote {target}")
    for title, cat, _n in SECTIONS:
        qs = bank[cat]
        print(f"  {title:<24}{len(qs):>3} questions   {qs[0]['q']}-{qs[-1]['q']}")
    print(f"  {'TOTAL':<24}{grand:>3} questions — verified identical to the live bank")


if __name__ == "__main__":
    if len(sys.argv) > 1:
        build(Path(sys.argv[1]).resolve())
    else:
        build(PROJECT / "docs" / NAME)
        try:
            import shutil
            shutil.copy2(PROJECT / "docs" / NAME, PROJECT.parent / NAME)
            print(f"mirrored to {PROJECT.parent / NAME}")
        except OSError as e:
            print(f"could not mirror: {e}")
