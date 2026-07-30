# -*- coding: utf-8 -*-
"""Rebuild "finalised set 2026.xlsx" from the client's updated 2026 questions.

Source of truth for QUESTION TEXT is the client's own 2026 workbook. This script
holds that content transcribed and cleaned, plus the gaps filled per the agreed
decisions:

  1. interests  — client supplied 10, blueprint needs 12; Q11 and Q12 authored
                  here in the client's own 5-family format.
  2. Strenghts  — the client sheet measured career interests (same Tech/Health/
                  Agri/Law buckets as `interests`, with each option labelled with
                  its bucket), not the 8 strength domains. Rebuilt: 8 new items
                  targeting Analytical, Creative, Leadership, Relationship,
                  Execution, Communication, Adaptability, Learning.
  3. Scoring    — old mappings are reused wherever the client's question still
                  matches the old construct and option order (motivators A-D,
                  learning-style VARK positions, EI per-question dimension,
                  personality trait/facet). Where the client rewrote the options
                  so the old weights no longer line up, the weight column is left
                  blank and marked "TO CONFIRM" rather than guessed at.

Old counts are preserved exactly (60 questions, Q1-Q60). Sheets where the client
supplied more than the blueprint count are cut to the top N — for aptitude, EI
and learning styles the selection is stated per row in the "Source" column.

Aptitude additionally carries Image / Image Specification columns and an explicit
Correct Answer column.

Usage:  python scripts/build_finalised_set_2026.py [output.xlsx]
"""

from __future__ import annotations

import shutil
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HERE = Path(__file__).resolve().parent
PROJECT = HERE.parent          # the git repo root (github.com/mahendrabella1/CareerFitment)
OUTER = PROJECT.parent         # the enclosing Desktop folder, NOT under version control

# Canonical, version-controlled output lives inside the repo so the question set
# ships with the code that scores it.
TARGET = PROJECT / "docs" / "finalised set 2026.xlsx"
# Convenience mirror in the enclosing folder, where the workbook is normally
# opened. Best-effort: skipped without complaint if Excel is holding it open.
MIRROR = OUTER / "finalised set 2026.xlsx"

TBD = "TO CONFIRM"

# --------------------------------------------------------------------------- #
# Blueprint — the old count. Nothing may exceed these.
# --------------------------------------------------------------------------- #
BLUEPRINT = [
    ("interests", 12, "Q1-Q12"),
    ("aptitude", 10, "Q13-Q22"),
    ("personality", 12, "Q23-Q34"),
    ("Strenghts", 8, "Q35-Q42"),
    ("motivators", 5, "Q43-Q47"),
    ("Learning styles", 4, "Q48-Q51"),
    ("multiple intellligence", 4, "Q52-Q55"),
    ("emotional intelligence", 5, "Q56-Q60"),
]

BLUEPRINT_NOTE = {
    "interests": "5 career families per item: Health · Agriculture/Science · Engineering-Tech · Law-Arts · Business",
    "aptitude": "10 distinct dimensions, one item each. 3 items need artwork, 1 recommended.",
    "personality": "Big Five — Openness, Conscientiousness, Extraversion, Agreeableness, Emotional Stability",
    "Strenghts": "8 strength domains — Analytical, Creative, Leadership, Relationship, Execution, Communication, Adaptability, Learning",
    "motivators": "Achievement, Innovation, Impact, Leadership, Security",
    "Learning styles": "VARK — Visual, Aural, Read/Write, Kinesthetic, Multimodal",
    "multiple intellligence": "Logical, Visual-Spatial, Linguistic/Interpersonal, Intrapersonal, Bodily-Kinesthetic",
    "emotional intelligence": "Self-Awareness, Self-Regulation, Self-Motivation, Empathy, Relationship Management",
}

CLIENT = "Client 2026 sheet"
AUTHORED = "Authored — please review"

# =========================================================================== #
# INTERESTS — client's 10 (cleaned out of the CSV paste) + 2 authored.
# Each option carries the career family the client assigned it, plus a RIASEC
# weight derived from what the option actually involves.
#
# SCORING RULE: read the cluster from the per-option tag, NOT from the option
# letter. Across the client's sheet, positions C and D alternate between Law,
# Engineering and Arts (Q1 C=Law but Q3 C=Engineering; Q2 D=Arts but Q8 D=Law),
# so positional scoring would mis-assign clusters. Tag-based scoring makes the
# inconsistency harmless and the client's wording can stay exactly as written.
# =========================================================================== #
INTERESTS = [
    ("Q1", "Your school is organizing a District level Community Fair. Which role would you naturally pick?", [
        ("Diagnose community health issues and suggest remedies.", "C", "Health & Medical / Allied Health: Doctor, Nurse, Physiotherapist, Pharmacist", {"S": 3, "I": 2}),
        ("Build a tech device, software, or mechanical solution.", "B", "Engineering & Tech: Software Developer, Mechanical Engineer, Robotics Engineer", {"R": 3, "I": 2}),
        ("Research local laws and policies, and present solutions to officials.", "F", "Law & Public Policy: Lawyer, Civil Servant (IAS), Policy Analyst, Mediator", {"E": 2, "S": 2, "I": 1}),
        ("Create posters, films, and branding for the fair.", "D", "Arts, Media & Design: Graphic Designer, Filmmaker, Journalist, Content Creator", {"A": 3, "E": 1}),
        ("Manage stall budgets, sponsorship money, and resource logistics.", "E", "Business, Finance & Mgmt: Chartered Accountant, Financial Analyst, Operations Mgr", {"C": 3, "E": 2}),
    ], CLIENT, ""),
    ("Q2", "If you could shadow an expert for a week during your summer break, who would you choose?", [
        ("A surgeon performing complex medical procedures or lab research.", "C", "Health & Medical: Surgeon, Medical Researcher, Pathologist, Biomedical Scientist", {"I": 3, "S": 2}),
        ("An agronomist working on organic farming and crop genetics.", "G", "Agriculture & Allied Sciences: Agronomist, Food Technologist, Environmental Scientist", {"R": 3, "I": 2}),
        ("A corporate lawyer negotiating major international business deals.", "F", "Law & Commerce: Corporate Lawyer, Legal Consultant, Compliance Officer", {"E": 3, "C": 1}),
        ("An animator or creative director producing a feature film.", "D", "Arts & Design: Animator, Game Designer, Fashion Designer, Art Director", {"A": 3}),
        ("A startup founder pitching to investors and building a company.", "E", "Business & Entrepreneurship: Founder/CEO, Venture Capitalist, Product Manager", {"E": 3, "A": 1}),
    ], CLIENT, ""),
    ("Q3", "Your town is setting up a new Model Village Project. Where would you make the biggest impact?", [
        ("Set up free health screening camps and wellness awareness.", "C", "Health & Allied Health: Community Health Specialist, Nutritionist, Public Health Mgr", {"S": 3, "I": 1}),
        ("Install modern irrigation, soil testing, and sustainable farming systems.", "G", "Agriculture & Earth Sciences: Agricultural Engineer, Soil Scientist, Forester", {"R": 3, "I": 1}),
        ("Construct smart bridges, renewable energy grids, and water systems.", "A", "Engineering: Civil Engineer, Renewable Energy Specialist, Electrical Engineer", {"R": 3, "I": 2}),
        ("Archive local history, write village stories, and preserve regional art.", "D", "Humanities & Social Sciences: Historian, Anthropologist, Writer, Sociologist", {"A": 3, "I": 1}),
        ("Set up micro-finance banks and manage project budgets.", "E", "Finance & Banking: Investment Banker, Microfinance Officer, Risk Analyst", {"C": 3, "E": 2}),
    ], CLIENT, ""),
    ("Q4", "For your Class 10 annual exhibition, which project topic would you be most eager to lead?", [
        ("Studying human anatomy, genetics, or disease prevention methods.", "C", "Medical & Life Sciences: Physician, Geneticist, Microbiologist, Biotechnologist", {"I": 3, "S": 1}),
        ("Building an automated solar tracker or smart home automation kit.", "B", "Engineering & AI: AI Engineer, Electronics Engineer, Energy Systems Engineer", {"R": 3, "I": 2}),
        ("Debating constitutional rights, international relations, or student laws.", "F", "Law, Politics & Humanities: Criminal Lawyer, Diplomat, Political Analyst, Advocate", {"E": 2, "S": 2}),
        ("Displaying fine arts, photography, digital illustration, or set design.", "D", "Arts & Creative Fields: Fine Artist, Photographer, Interior Designer, Illustrator", {"A": 3}),
        ("Running a live stock-market simulation or business pitch deck.", "E", "Business & Economics: Economist, Stock Broker, Marketing Manager, Business Analyst", {"E": 3, "C": 2}),
    ], CLIENT, ""),
    ("Q5", "If you were given ₹10,000 to launch a student-led initiative, what would you fund?", [
        ("Rehabilitation equipment or mental health therapy sessions for peers.", "C", "Allied Health & Psychology: Clinical Psychologist, Occupational Therapist, Speech Therapist", {"S": 3, "I": 1}),
        ("Vertical hydroponic kits or organic seed beds for your school garden.", "G", "Agriculture & Food Tech: Food Scientist, Horticulturist, Agricultural Businessman", {"R": 3, "I": 1}),
        ("A mobile coding lab or 3D printing setup for young inventors.", "B", "Tech & Hardware: Hardware Engineer, Software Architect, Mechatronics Specialist", {"I": 3, "R": 2}),
        ("A theatre production, podcast studio, or school newspaper magazine.", "D", "Media, Journalism & Arts: Journalist, Theatre Director, Radio Jockey, Copywriter", {"A": 3, "E": 1}),
        ("An e-commerce store reselling handmade goods to earn profit.", "E", "Commerce & Management: E-commerce Manager, Sales Director, Financial Planner", {"E": 3, "C": 2}),
    ], CLIENT, ""),
    ("Q6", "Which type of books, documentary channels, or podcasts capture your attention most?", [
        ("Medical breakthroughs, human brain mysteries, or emergency ER stories.", "C", "Health & Medical Sciences: Neuroscientist, Doctor, Radiologist, Anesthetist", {"I": 3, "S": 1}),
        ("Wildlife conservation, forest ecosystems, and sustainable agriculture.", "G", "Environment & Agriculture: Wildlife Biologist, Zoologist, Environmental Lawyer", {"I": 3, "R": 2}),
        ("Space exploration, coding tutorials, and advanced robotics breakthroughs.", "B", "Science & Technology: Astrophysicist, Data Scientist, Cybersecurity Analyst", {"I": 3, "R": 2}),
        ("Crime thrillers, courtroom dramas, philosophy, and history podcasts.", "F", "Law, Criminology & Arts: Criminologist, Judge, Legal Journalist, Philosopher", {"A": 2, "S": 1, "E": 1}),
        ("Case studies on Fortune 500 companies, stock markets, and economics.", "E", "Business & Finance: Business Consultant, Chartered Accountant (CA), CFA, Auditor", {"E": 3, "C": 2}),
    ], CLIENT, ""),
    ("Q7", "During a crisis like a sudden epidemic in a school hostel, what is your instinct?", [
        ("Administer first aid, isolate the sick, and monitor physical symptoms.", "C", "Medical & Emergency Health: Emergency Doctor, Epidemiologist, Paramedic", {"S": 3, "R": 1}),
        ("Comfort anxious students, offer counseling, and boost team morale.", "F", "Psychology & Human Resources: Counselor, Psychologist, HR Manager, Social Worker", {"S": 3, "A": 1}),
        ("Trace how the infection is spreading — food, water or contact — and test the samples.", "G", "Science & Public Health: Microbiologist, Epidemiologist, Lab Scientist", {"I": 3, "R": 2}),
        ("Create clear infographics and announcements to keep everyone informed.", "D", "Design & Communication: PR Specialist, Communications Mgr, Visual Designer", {"A": 3, "E": 1}),
        ("Audit food and water supplies, manage logistics, and secure needed funds.", "E", "Operations & Supply Chain: Supply Chain Manager, Procurement Officer, Logistics Head", {"C": 3, "E": 2}),
    ], CLIENT + " (option C replaced)",
     "CHANGED. The client's option C ('Study health guidelines, ensure legal protocols, enforce safety rules' — "
     "Law & Administration) mapped to Human & Public Services, the same cluster as option B, so two of the five "
     "choices gave an identical signal and no Science/Agriculture option was offered. Replaced with a "
     "laboratory/epidemiology option so all five families appear once. Revert from the client-original file if "
     "you prefer the legal-protocols wording."),
    ("Q8", "What kind of practical problem would you feel most proud to solve in your career?", [
        ("Finding affordable cures or treatments for rare diseases.", "C", "Medical Research: Pharmacologist, Oncologist, Genetic Researcher", {"I": 3, "S": 2}),
        ("Improving crop yields to tackle hunger without harming the soil.", "G", "Agriculture Science: Soil Chemist, Plant Breeder, Agricultural Economist", {"R": 3, "I": 2}),
        ("Designing cleaner engines, faster computers, or automated systems.", "A", "Core Engineering: Automobile Engineer, Computer Engineer, Chemical Engineer", {"R": 3, "I": 2}),
        ("Defending human rights, fighting injustice, or reforming laws.", "F", "Legal & Social Services: Human Rights Lawyer, Judge, NGO Leader, Policy Maker", {"S": 3, "E": 2}),
        ("Restructuring a struggling business to make it highly profitable.", "E", "Corporate Management: Management Consultant, Chief Financial Officer (CFO), Strategist", {"E": 3, "C": 2}),
    ], CLIENT, "Client text read 'an struggling business' — corrected to 'a struggling business'."),
    ("Q9", "What kind of work environment sounds most appealing to you long-term?", [
        ("A modern hospital, clinical lab, or emergency medical center.", "C", "Health Sciences: Physician, Clinical Researcher, Surgeon, Medical Technologist", {"S": 3, "I": 2}),
        ("Outdoor fields, greenhouses, research farms, or natural reserves.", "G", "Agriculture & Earth: Forester, Environmental Consultant, Agricultural Scientist", {"R": 3, "I": 1}),
        ("A modern tech hub, engineering workshop, or R&D lab.", "B", "Engineering & Tech: Software Engineer, Mechanical Engineer, Robotics Researcher", {"R": 3, "I": 3}),
        ("A newsroom, media studio, or publishing house.", "D", "Media & Creative Arts: Journalist, Editor, Creative Director, Producer", {"A": 3, "E": 1}),
        ("A corporate boardroom, stock exchange floor, or financial firm.", "E", "Business & Finance: Investment Banker, Equity Analyst, Corporate Executive", {"E": 3, "C": 3}),
    ], CLIENT + " (option D narrowed)",
     "CHANGED. The client's option D read 'A courtroom, law firm, media studio, or publishing house', mixing Law "
     "and Creative Arts in one choice, so a student picking it was ambiguous between clusters F and D. Narrowed "
     "to the media/publishing half. Law is already covered on Q1, Q2, Q4, Q6, Q8 and Q10."),
    ("Q10", "If you could take a specialized elective course next term, which would you pick?", [
        ("Human Physiology, Nutrition, and Clinical Health Basics.", "C", "Health & Allied Sciences: Dietitian, Physiotherapist, Sports Medicine Specialist", {"I": 3, "S": 2}),
        ("Environmental Science, Crop Care, and Biotechnology.", "G", "Life Sciences & Agriculture: Botanist, Biotechnologist, Agronomist", {"I": 3, "R": 2}),
        ("Python Programming, Electronics, and Applied Mathematics.", "B", "Engineering & Analytics: Data Analyst, Software Engineer, Electrical Engineer", {"I": 3, "R": 2}),
        ("World History, Constitutional Law, and Creative Writing.", "F", "Law & Humanities: Historian, Political Scientist, Journalist, Legal Scholar", {"A": 3, "S": 1}),
        ("Financial Accounting, Entrepreneurship, and Business Marketing.", "E", "Commerce & Management: Accountant, Marketing Strategist, Finance Manager", {"C": 3, "E": 3}),
    ], CLIENT, ""),
    ("Q11", "Your school has one slot left in the district science-and-innovation expo. Which entry would you put your name to?", [
        ("A low-cost device that screens for anaemia without taking blood.", "C", "Health & Medical: Biomedical Engineer, Pathologist, Public Health Researcher", {"I": 3, "S": 2}),
        ("A soil-health sensor that tells farmers exactly what their field is missing.", "G", "Agriculture & Earth Sciences: Soil Scientist, Agricultural Engineer, Agronomist", {"R": 3, "I": 2}),
        ("An app that maps and reports unsafe road stretches around the school.", "B", "Engineering & Tech: Software Developer, Data Analyst, Civil Engineer", {"I": 3, "R": 1, "S": 1}),
        ("A short documentary on a local craft that is disappearing.", "D", "Arts, Media & Design: Filmmaker, Photographer, Journalist", {"A": 3}),
        ("A costed business plan to take a student product to real customers.", "E", "Business, Finance & Mgmt: Entrepreneur, Financial Analyst, Marketing Manager", {"E": 3, "C": 2}),
    ], AUTHORED, "Authored to bring interests up to the blueprint count of 12. Follows the client's 5-family format."),
    ("Q12", "Ten years from now a magazine profiles you. Which headline would you be happiest to read?", [
        ("“The doctor who brought specialist care to villages that had none.”", "C", "Health & Medical: Physician, Public Health Specialist, Surgeon", {"S": 3, "I": 2}),
        ("“The scientist who made a drought-hit district farm again.”", "G", "Agriculture & Earth Sciences: Agricultural Scientist, Botanist, Environmental Scientist", {"I": 3, "R": 2}),
        ("“The engineer behind the system half the country now uses.”", "B", "Engineering & Tech: Software Architect, Systems Engineer, AI Engineer", {"I": 3, "R": 2}),
        ("“The film-maker whose first feature changed the conversation.”", "D", "Arts, Media & Design: Filmmaker, Creative Director, Writer", {"A": 3}),
        ("“The founder who built a company out of a school project.”", "E", "Business & Entrepreneurship: Founder/CEO, Business Strategist, Investor", {"E": 3, "C": 1}),
    ], AUTHORED, "Authored to bring interests up to the blueprint count of 12. Follows the client's 5-family format."),
]

INTERESTS_SCORING_NOTE = (
    "SCORING RULE — read each option's cluster from its own tag in the 'Career Cluster (per option)' column, not "
    "from the option letter. Across this sheet the client's option positions are not consistent: position C is Law "
    "on Q1/Q2/Q4 but Engineering on Q3/Q8, and position D is Arts on Q2/Q4/Q5 but Law on Q6/Q8/Q10. Tag-based "
    "scoring makes that harmless; positional scoring would mis-assign clusters on roughly half the items."
)

# =========================================================================== #
# APTITUDE — 10 of the client's 15, chosen to give one item per dimension.
# (question, options A-E, correct letter, dimension, clusters, professions,
#  image flag, image spec, source, review note)
# =========================================================================== #
APTITUDE = [
    ("Q13", "A student business buys notebooks at ₹40 each and sells them for ₹50 each. What is the profit percentage?",
     ["15%", "20%", "25%", "30%", "10%"], "C",
     "Numerical Reasoning", "Engineering, Finance, Analytics", "Engineer, Chartered Accountant, Data Scientist, Economist",
     "Not needed", "Text-only item.", CLIENT + " (item 11)",
     "Retagged. Client tag was 'Business, Finance & Law / CA, Financial Analyst, Investment Banker' — kept the "
     "finance thrust but widened to the clusters numerical reasoning actually predicts."),
    ("Q14", "All doctors study science. Some science students become researchers. Which conclusion is definitely true?",
     ["All researchers are doctors.", "All science students become doctors.",
      "Some people who study science become researchers.", "No doctors are researchers.",
      "All researchers are science teachers."], "C",
     "Logical Deduction", "IT, Engineering, Research, Law", "Software Engineer, Lawyer, Research Scientist, AI Engineer",
     "Not needed", "Text-only item.", CLIENT + " (item 12)",
     "Retagged from 'Business, Finance & Law' — syllogistic reasoning predicts IT/research as strongly as law."),
    ("Q15", "Which of the following items is the odd one out among specialized professional tools?",
     ["Python (Programming)", "Stethoscope", "Ledger", "Palette", "Tractor"], "E",
     "Verbal Classification", "Education, Law, Media", "Teacher, Lawyer, Journalist, Editor",
     "Optional", "Supporting illustration: five small line icons, one per option (laptop, stethoscope, ledger, palette, tractor).",
     CLIENT + " (item 13)",
     "Retagged. Client tag was 'Society, Nature & Public Service / Agricultural Engineer, Farm Operator', which "
     "was read off the correct answer (Tractor) rather than the skill. Verbal classification predicts "
     "language-heavy fields."),
    ("Q16", "School electricity usage dropped from 1,200 kWh in January to 900 kWh in February. What is the percentage reduction?",
     ["20%", "25%", "30%", "33.3%", "15%"], "B",
     "Data Interpretation", "Finance, AI, Business Analytics", "Data Analyst, Actuary, Economist, Business Analyst",
     "Recommended", "Two-bar column chart — January 1,200 kWh and February 900 kWh, y-axis labelled kWh. Needed so "
     "the student reads the figures off a display rather than out of prose, which is what Data Interpretation means.",
     CLIENT + " (item 14)",
     "Retagged from 'Health & Life Sciences / Environmental Health Specialist, Bio-Statistician', which did not "
     "follow from an electricity-usage item. Chart still to be drawn."),
    ("Q17", "A square piece of paper is folded twice and a single triangular punch is made through all the layers. When the paper is unfolded, how many punched holes appear?",
     ["2", "4", "6", "8", "1"], "B",
     "Spatial Reasoning", "Architecture, Civil, Mechanical", "Architect, Civil Engineer, Pilot, Surgeon",
     "Required", "Four-panel strip: (1) plain square, (2) folded in half, (3) folded in half again, "
     "(4) the folded square with a triangular notch cut from one edge. Do not show the unfolded result.",
     CLIENT + " (item 15)",
     "Client marked this '(visual item)' — artwork is mandatory, the item cannot be answered without it. "
     "Retagged from 'Health & Life Sciences / Orthopedic Specialist', which did not follow from paper folding."),
    ("Q18", "Diagnosis is to Treatment as Evidence is to ______.",
     ["Investigation", "Medicine", "Verdict", "Crime", "Policy"], "C",
     "Verbal Analogy", "Law, Education, Media", "Lawyer, Teacher, Journalist, Psychologist",
     "Not needed", "Text-only item.", CLIENT + " (item 17)",
     "Retagged from 'Business, Finance & Law' — analogy reasoning is a language aptitude, so education and media "
     "belong alongside law."),
    ("Q19", "If TECH is written as 3142 and ARTS is written as 5637, how is RATE written?",
     ["6513", "6531", "5631", "6351", "3561"], "B",
     "Coding-Decoding", "IT, Cryptography, Research", "Software Engineer, Cybersecurity Analyst, Data Scientist",
     "Not needed", "Text-only item.", CLIENT + " (item 18, corrected)",
     "FIXED. The client keys contradicted each other: TECH=3142 gives T=3 but ARTS=5678 gives T=7, and the marked "
     "answer 6571 silently assumed T=7. Second key changed to ARTS=5637 so T=3 throughout (A=5, R=6, T=3, S=7), "
     "making RATE = 6531 the single defensible answer. Options rewritten around it; correct answer stays at B. "
     "Also retagged from 'Health & Life Sciences / Genetic Coder'."),
    ("Q20", "A series shows 3 triangles, then 5 squares, then 7 pentagons. What shape and count come next?",
     ["8 Hexagons", "9 Pentagons", "9 Hexagons", "10 Heptagons", "11 Hexagons"], "C",
     "Abstract Reasoning", "Design, AI, Innovation", "Architect, UX Designer, Product Designer, AI Engineer",
     "Required", "Three groups drawn left to right — 3 triangles, 5 squares, 7 pentagons — then a boxed '?'. "
     "Counts must be visibly correct so the 3/5/7 progression can be seen, not just read.",
     CLIENT + " (item 19, corrected)",
     "FIXED. The client series started at circles (3 circles, 5 triangles, 7 squares), so the two rules "
     "disagreed — count went +2 but the side-count rule broke at circle-to-triangle, leaving D arguable. "
     "Series now starts at a triangle: counts +2 and sides +1 both hold cleanly, giving 9 hexagons. Correct "
     "answer stays at C. Also retagged from 'Business, Finance & Law'."),
    ("Q21", "Which number is different?",
     ["456789", "456789", "456879", "456789", "All are the same"], "C",
     "Attention to Detail", "Healthcare, Finance, Quality", "Doctor, Pharmacist, Auditor, QA Engineer",
     "Optional", "Supporting illustration: the five choices set in a monospaced vertical stack so digits align column by column.",
     CLIENT + " (item 14, corrected)",
     "FIXED. The client stem listed 456789, 456798, 456879, 456789 — because 456789 appeared twice, both 456798 "
     "and 456879 were unique, so the item had two correct answers. Option E was also the stray letter 'e'. "
     "Rewritten to the old set's pattern: three identical values, one different, plus an 'All are the same' distractor."),
    ("Q22", "Two gears mesh together. Gear A turns clockwise. Which way does Gear B turn?",
     ["Clockwise", "Counter-clockwise", "No movement", "Depends on size", "Cannot say"], "B",
     "Mechanical Reasoning", "Mechanical, Robotics, Defence", "Mechanical Engineer, Robotics Engineer, Aerospace Engineer",
     "Required", "Two meshed spur gears side by side, labelled 'Gear A' and 'Gear B'. A curved arrow on Gear A "
     "marked 'clockwise'; a '?' above Gear B. This artwork already exists as inline SVG in the old aptitude bank "
     "(9-10 Set 1, Q22) and can be reused as-is.",
     CLIENT + " (item 15)", ""),
]

APTITUDE_DROPPED = [
    ("item 16 — bicycle at 12 km/h for 30 minutes", "Applied Math duplicates Numerical Reasoning (already covered by Q13)."),
    ("item 20 — drip irrigation, direct cause", "Statement & Cause; good item, cut only because the blueprint caps aptitude at 10."),
    ("block 2 item 11 — wildlife reserve, 480 animals", "Numerical Reasoning duplicate."),
    ("block 2 item 12 — business earns ₹25,000, 20% more", "Financial Reasoning duplicate of Q13's profit calculation."),
    ("block 2 item 13 — football team, 3 goals per match", "Numerical Reasoning duplicate."),
]

# =========================================================================== #
# PERSONALITY — the client's 12 scenarios and their 12 x 5 choices, with the
# option shorthand expanded into student-facing sentences (meaning unchanged;
# the client's exact wording is preserved in the client-original workbook).
# Trait and facet carried over from the old set; per-option Big Five weights
# derived from the client's option meanings using the old set's convention
# (primary trait +3, supporting traits +1/+2).
# O = Openness · C = Conscientiousness · E = Extraversion
# A = Agreeableness · S = Emotional Stability
# =========================================================================== #
PERSONALITY = [
    ("Q23", "You have one weekend to prepare a project on an unfamiliar topic. By Saturday evening you still have several possible directions. What would you most likely do?",
     [("Go with the most unusual idea, even if it might not work.", {"O": 3, "S": 1}),
      ("Go with the clearest, most reliable approach.", {"C": 3, "S": 1}),
      ("Talk the options through with classmates before choosing.", {"E": 2, "A": 2}),
      ("Combine the best parts of several different sources.", {"O": 3, "C": 1}),
      ("None of these.", {})],
     "Openness", "Curiosity",
     "Client option B read 'Clearest, idea most reliable approach' — the stray 'idea' looked like a typo and has "
     "been dropped."),
    ("Q24", "Your school is holding an exhibition and you must choose one project. Which would you most enjoy building?",
     [("A project that challenges how people normally think.", {"O": 3}),
      ("A carefully planned project that I have tested properly.", {"C": 3}),
      ("An interactive project that visitors can take part in.", {"E": 3, "A": 2}),
      ("A visually creative project that people will remember.", {"O": 3, "A": 1}),
      ("None of these.", {})],
     "Openness", "Creativity", ""),
    ("Q25", "An assignment is due in two weeks, but an unexpected event cuts down your study days. What would you most likely do?",
     [("Adjust my schedule and follow the revised plan.", {"C": 3, "S": 1}),
      ("Rethink the best approach before starting again.", {"O": 2, "C": 1}),
      ("Work longer hours to catch up.", {"C": 2, "S": 1}),
      ("Study with others to stay on track.", {"E": 2, "A": 2}),
      ("None of these.", {})],
     "Conscientiousness", "Planning", ""),
    ("Q26", "After an hour studying a hard chapter you still don't understand it. What usually happens next?",
     [("I try a different way of studying it.", {"O": 2, "C": 1}),
      ("I take a break and come back to it later.", {"S": 2}),
      ("I ask someone to explain it to me.", {"E": 2, "A": 1}),
      ("I keep practising even if the theory is still unclear.", {"C": 3}),
      ("None of these.", {})],
     "Conscientiousness", "Persistence", ""),
    ("Q27", "Halfway through organizing a school event, several problems appear at once. Which role do you take?",
     [("Keeping everything organised and on track.", {"C": 3}),
      ("Suggesting new ideas and solutions.", {"O": 3}),
      ("Keeping everyone coordinated and motivated.", {"E": 3, "A": 2}),
      ("Dealing with the most urgent problem first.", {"S": 3, "C": 1}),
      ("None of these.", {})],
     "Conscientiousness", "Responsibility", ""),
    ("Q28", "You arrive early for a workshop and most students are already talking. What would you most likely do?",
     [("Join the conversation naturally.", {"E": 3, "A": 1}),
      ("Watch for a while before joining in.", {"C": 1, "O": 1}),
      ("Start talking once I notice a shared interest.", {"E": 2, "A": 1}),
      ("Wait until someone includes me.", {}),
      ("None of these.", {})],
     "Extraversion", "Initiative",
     "Facet changed from the old set's 'Confidence'. The old Q28 measured confidence through public speaking; "
     "this scenario measures whether the student initiates socially, which is distinct from Q29's question about "
     "how the first week turned out. Calling it Initiative keeps Q28 and Q29 from measuring the same thing."),
    ("Q29", "You joined a new class for a semester. By the end of the first week, what is most likely true?",
     [("I have already introduced myself to a lot of people.", {"E": 3, "A": 2}),
      ("I have spoken mainly when I needed to.", {"C": 2}),
      ("I have connected with a few people like me.", {"E": 2, "A": 2}),
      ("I have focused on settling in first.", {"C": 2}),
      ("None of these.", {})],
     "Extraversion", "Social Interaction", ""),
    ("Q30", "During a team activity, two members disagree on how to complete the work. What would you naturally do?",
     [("Help both sides understand each other.", {"A": 3, "S": 1}),
      ("Combine the strongest parts of both ideas.", {"A": 2, "O": 2}),
      ("Let the group decide and focus on my own task.", {"C": 2}),
      ("Choose whichever option is most practical.", {"C": 2, "S": 1}),
      ("None of these.", {})],
     "Agreeableness", "Cooperation",
     "Facets swapped relative to the old set: old Q30 measured Empathy and old Q31 Cooperation. Mapped to match "
     "the client's new scenarios, which are the other way round."),
    ("Q31", "A teammate is falling behind and the deadline is approaching. What feels most natural?",
     [("Support them, but encourage them to do their own part.", {"A": 3, "S": 1}),
      ("Adjust the plan so everyone contributes fairly.", {"A": 2, "C": 2}),
      ("Finish part of their task to help the team.", {"A": 2, "C": 2}),
      ("Focus on my own responsibilities first.", {"C": 2}),
      ("None of these.", {})],
     "Agreeableness", "Empathy",
     "Facets swapped relative to the old set — see Q30."),
    ("Q32", "Your result is much lower than expected after working hard. What would you most likely do next?",
     [("Review my mistakes and adjust my approach.", {"S": 3, "C": 2}),
      ("Ask for feedback on where I went wrong.", {"S": 2, "A": 1, "E": 1}),
      ("Take a short break, then start again.", {"S": 2}),
      ("Accept the result and keep working steadily.", {"S": 1, "C": 1}),
      ("None of these.", {})],
     "Emotional Stability", "Emotional Control", ""),
    ("Q33", "Just before presenting, you realize an important part is missing. What is your first reaction?",
     [("Quickly find another way to explain it.", {"S": 3, "O": 1}),
      ("Pause, understand the situation, then continue.", {"S": 3, "C": 2}),
      ("Ask someone nearby for help.", {"A": 2, "E": 2, "S": 1}),
      ("Carry on and adapt as I go.", {"S": 2, "C": 1}),
      ("None of these.", {})],
     "Emotional Stability", "Composure", ""),
    ("Q34", "You have four internship offers with the same salary and growth. Which would you most enjoy?",
     [("One where I learn something completely new.", {"O": 3, "C": 2}),
      ("One where I create original ideas, designs or solutions.", {"O": 3, "E": 1}),
      ("One where I work closely with people and help teams.", {"E": 3, "A": 3}),
      ("One where I improve systems and boost efficiency.", {"C": 3, "S": 1}),
      ("None of these.", {})],
     "Integrated Big Five", "Overall Personality Profile", ""),
]

PERSONALITY_GLOBAL_NOTES = (
    "SCORING RULE for option E. 'None of these' is non-scoring by design on all 12 items. Do NOT divide each "
    "trait by 12 — normalise it over the number of items the student actually scored on, so a student who picks E "
    "a few times still gets a valid profile instead of an artificially flattened one. If a student picks E on "
    "more than half the items, suppress the Big Five profile and report it as incomplete rather than publishing a "
    "near-zero result. Note also that option texts were expanded from the client's shorthand into full sentences "
    "for Class 9-10 readability; the original wording is preserved in 'finalised set 2026 (client original).xlsx'."
)

# =========================================================================== #
# STRENGHTS — rebuilt to measure the 8 strength domains (client sheet measured
# career interests instead). Each option names its domain + weights.
# =========================================================================== #
STRENGTHS = [
    ("Q35", "Your team has six weeks to build an entry for a national school challenge. In the first week, what are you doing?", [
        ("Breaking the brief down and working out what actually wins marks.", {"Analytical": 3, "Learning": 1}),
        ("Sketching an idea nobody else will have thought of.", {"Creative": 3, "Analytical": 1}),
        ("Sorting out who does what, and making sure nobody drops out.", {"Leadership": 3, "Relationship": 1}),
        ("Turning the deadline into a week-by-week plan and starting task one.", {"Execution": 3, "Analytical": 1}),
        ("Reading up on the subject until I know it better than the rest of the team.", {"Learning": 3, "Analytical": 1}),
    ]),
    ("Q36", "Your class presents to parents tomorrow and the presentation is half finished. What do you take on?", [
        ("Checking every fact and figure before it goes on a slide.", {"Analytical": 3, "Execution": 1}),
        ("Finding an angle that will actually hold the room.", {"Creative": 3, "Communication": 1}),
        ("Standing up and delivering it.", {"Communication": 3, "Leadership": 1}),
        ("Building the slides and fixing the running order.", {"Execution": 3, "Creative": 1}),
        ("Keeping everyone calm and working until it is done.", {"Relationship": 3, "Leadership": 1}),
    ]),
    ("Q37", "A new student joins mid-term and is clearly struggling to settle in. What do you do?", [
        ("Work out what is actually making it hard for them before doing anything.", {"Analytical": 3, "Learning": 1}),
        ("Introduce them around and sit with them for a few days.", {"Relationship": 3, "Communication": 1}),
        ("Raise it with the class so everyone makes an effort.", {"Communication": 3, "Leadership": 1}),
        ("Set up a buddy rota and make sure it keeps running.", {"Execution": 3, "Relationship": 1}),
        ("Get them into an activity they would be good at, so people see them differently.", {"Creative": 3, "Relationship": 1}),
    ]),
    ("Q38", "Four weeks into a project you realise the approach you chose will not work. What is your first move?", [
        ("Trace back to exactly where it started going wrong.", {"Analytical": 3, "Execution": 1}),
        ("Switch to a different approach and keep moving.", {"Adaptability": 3, "Execution": 1}),
        ("Look up how other people have solved something similar.", {"Learning": 3, "Analytical": 1}),
        ("Get the team together and rethink it with them.", {"Relationship": 3, "Communication": 1}),
        ("Rebuild the schedule around a method I know will finish on time.", {"Execution": 3, "Leadership": 1}),
    ]),
    ("Q39", "Your school gives your group a budget and complete freedom for one project. Which role do you want?", [
        ("Deciding what we do, and holding us to it.", {"Leadership": 3, "Execution": 1}),
        ("Coming up with the idea itself.", {"Creative": 3, "Learning": 1}),
        ("Pitching it to the school and to sponsors.", {"Communication": 3, "Leadership": 1}),
        ("Running the budget, the timeline and the paperwork.", {"Execution": 3, "Analytical": 1}),
        ("Learning whatever new skill the project turns out to need.", {"Learning": 3, "Adaptability": 1}),
    ]),
    ("Q40", "Two people in your group have stopped speaking to each other and the work has stalled. What do you do?", [
        ("Talk to each of them separately first.", {"Relationship": 3, "Communication": 1}),
        ("Get it out in the open with the whole group.", {"Communication": 3, "Leadership": 1}),
        ("Make a call on the disputed point so we can move on.", {"Leadership": 3, "Execution": 1}),
        ("Work out which of the two options is actually better on the evidence.", {"Analytical": 3, "Learning": 1}),
        ("Reshuffle the tasks so they do not have to work together.", {"Adaptability": 3, "Execution": 1}),
    ]),
    ("Q41", "You are handed something broken — an old radio, a bicycle, a piece of code — and told to make it work. How do you start?", [
        ("Take it apart methodically and find the fault.", {"Analytical": 3, "Execution": 1}),
        ("Read up on how it is supposed to work first.", {"Learning": 3, "Analytical": 1}),
        ("See whether it could become something better than it was.", {"Creative": 3, "Learning": 1}),
        ("Start trying things and adjust as I learn what breaks.", {"Adaptability": 3, "Creative": 1}),
        ("Break it into stages, get the parts I need, and work through to a finish.", {"Execution": 3, "Analytical": 1}),
    ]),
    ("Q42", "Looking back at the projects you have enjoyed the most, what usually made them enjoyable?", [
        ("Solving a hard problem and finding the answer.", {"Analytical": 3, "Learning": 2}),
        ("Making something original.", {"Creative": 3, "Adaptability": 1}),
        ("Working with people and getting there together.", {"Relationship": 3, "Leadership": 2}),
        ("Running it well and having it all come off.", {"Execution": 3, "Leadership": 1}),
        ("Getting genuinely good at something new.", {"Learning": 3, "Adaptability": 1}),
    ]),
]

STRENGTHS_CAREERS = {
    "Q35": ("Engineering & Technology, Business & Management, Design & Innovation, Research", "Software Engineer, Scientist, Entrepreneur, Architect, Product Manager"),
    "Q36": ("Marketing & Media, Education, Human Resources, Public Relations", "Teacher, Marketing Manager, Journalist, HR Manager, PR Specialist"),
    "Q37": ("Education, Human Resources, Psychology, Public Service", "Teacher, Counsellor, HR Manager, Social Worker, Psychologist"),
    "Q38": ("Research, Consulting, Healthcare, Education", "Doctor, Consultant, Teacher, Research Scientist, Psychologist"),
    "Q39": ("Business & Entrepreneurship, Marketing, Management", "Entrepreneur, Product Manager, Business Analyst, Marketing Manager"),
    "Q40": ("Human Resources, Management, Law, Public Administration", "HR Manager, Project Manager, Lawyer, IAS Officer, Team Leader"),
    "Q41": ("Core Engineering, Innovation, Research & Development", "Mechanical Engineer, Product Designer, AI Engineer, Innovation Consultant"),
    "Q42": ("Integrated validation — all career clusters", "Engineering, Healthcare, Business, Design, Research, Education, Government, Media, Finance"),
}

# =========================================================================== #
# MOTIVATORS — client's 5 (all five match the old constructs and option order,
# so the old A-D weights carry across). Option E authored.
# =========================================================================== #
MOTIVATORS = [
    ("Q43", "School gives you the opportunity to select one special project for the entire year. Which would excite you most?", [
        ("Challenging project to improve skills and achieve excellent results", {"Achievement": 3, "Innovation": 1, "Learning": 1}, CLIENT),
        ("Create something completely new nobody has tried", {"Innovation": 3, "Achievement": 1, "Impact": 1}, CLIENT),
        ("Solve a problem faced by people or society", {"Impact": 3, "Innovation": 1, "Leadership": 1}, CLIENT),
        ("Lead a team and make key decisions to achieve the goal", {"Leadership": 3, "Achievement": 1, "Impact": 1}, CLIENT),
        ("Take on a well-defined project with clear steps and a dependable result", {"Security": 3, "Achievement": 1}, AUTHORED),
    ]),
    ("Q44", "Extra time after school. Which activity would feel most satisfying?", [
        ("Practicing a skill to get better at it", {"Achievement": 3, "Learning": 2}, CLIENT),
        ("Exploring new ideas, experimenting creatively", {"Innovation": 3, "Learning": 1}, CLIENT),
        ("Helping someone learn or solve a problem", {"Impact": 3, "Learning": 1}, CLIENT),
        ("Organizing a group activity", {"Leadership": 3, "Achievement": 1}, CLIENT),
        ("Getting ahead on work already scheduled, so nothing piles up", {"Security": 3, "Achievement": 1}, AUTHORED),
    ]),
    ("Q45", "Achieved success in your future career. What would make you most proud?", [
        ("Becoming highly skilled, recognized as an expert", {"Achievement": 3, "Learning": 2}, CLIENT),
        ("Creating something innovative that changes how people do things", {"Innovation": 3, "Achievement": 1}, CLIENT),
        ("Making a meaningful difference in people's lives", {"Impact": 3}, CLIENT),
        ("Becoming a leader who influences important decisions", {"Leadership": 3, "Achievement": 1}, CLIENT),
        ("Building a secure, stable career that comfortably supports my family", {"Security": 3, "Impact": 1}, AUTHORED),
    ]),
    ("Q46", "School competition where students choose their own challenge. Which would you prefer?", [
        ("Difficult challenge to test abilities, compete with others", {"Achievement": 3, "Learning": 1}, CLIENT),
        ("Creative challenge to design something unique", {"Innovation": 3}, CLIENT),
        ("Challenge that helps improve school or community", {"Impact": 3}, CLIENT),
        ("Challenge to coordinate and manage a team", {"Leadership": 3}, CLIENT),
        ("A challenge with clear rules, where I know exactly what is expected", {"Security": 3, "Achievement": 1}, AUTHORED),
    ]),
    ("Q47", "Choosing your future workplace — which environment would motivate you most?", [
        ("Continuously learn, improve, achieve bigger goals", {"Learning": 3, "Achievement": 2}, CLIENT),
        ("Freedom to explore ideas and try new approaches", {"Innovation": 3, "Learning": 1}, CLIENT),
        ("Work that helps people, creates positive change", {"Impact": 3}, CLIENT),
        ("Stability, clear systems, structured career path", {"Security": 3}, CLIENT),
        ("A place where I can take charge of a team and be judged on results", {"Leadership": 3, "Achievement": 2}, AUTHORED),
    ]),
]

MOTIVATOR_CAREERS = {
    "Q43": ("Engineering, Research, Design, Healthcare, Entrepreneurship, Management", "Engineer, Scientist, Doctor, Entrepreneur, Project Manager"),
    "Q44": ("Technology, Education, Healthcare, Business, Creative Fields", "Software Engineer, Teacher, Doctor, Designer, Business Leader"),
    "Q45": ("Research, Technology, Healthcare, Social Services, Government", "Research Scientist, AI Engineer, Doctor, IAS Officer, Entrepreneur"),
    "Q46": ("Sports, Engineering, Design, Social Impact, Management", "Athlete, Engineer, Designer, Social Worker, Manager"),
    "Q47": ("Research, Entrepreneurship, Healthcare, Education, Finance, Government", "Scientist, Startup Founder, Doctor, Teacher, Banker, Civil Servant"),
}

# =========================================================================== #
# LEARNING STYLES — 4 of the client's 11, chosen to match the old four
# constructs. VARK positions come from the client's own column headers.
# =========================================================================== #
LEARNING = [
    ("Q48", "When learning a hard math formula, I like to",
     ["See charts/shapes", "Listen to explanation", "Write down step by step to understand",
      "Practice problems until I understand", "Use a mix — see it, talk it through, then practise it"],
     "Engineering, Research, Healthcare, Education, Design", "Engineer, Scientist, Doctor, Teacher, Designer",
     CLIENT + " (item 2) — matches old Q48 'difficult new concept'"),
    ("Q49", "When studying for a big final exam, I prefer to:",
     ["Look at colourful mind-maps or diagrams", "Join a study group to discuss topics out loud",
      "Reread chapters and write summaries", "Use hands-on practice methods",
      "Combine diagrams, discussion, notes and practice depending on the subject"],
     "All career clusters — study-approach indicator", "Supports learning approach in Engineering, Medicine, Business, Arts, Research",
     CLIENT + " (item 10) — matches old Q49 'exam preparation'"),
    ("Q50", "When learning a new hobby, I prefer to:",
     ["Watch video guides", "Have someone explain it", "Read online articles on it",
      "Try it by doing", "Switch between watching, asking, reading and trying as I go"],
     "Technology, Creative Fields, Entrepreneurship, Technical Careers", "Programmer, Designer, Entrepreneur, Engineer, Technician",
     CLIENT + " (item 7) — matches old Q50 'new skill outside academics'"),
    ("Q51", "You are exploring different career options for your future. How would you prefer to learn about them?",
     ["Watch videos of professionals at work", "Attend talks and meet counsellors directly",
      "Read guides and articles about professions", "Visit workplaces or try internships hands-on",
      "Use all of them — videos, talks, reading and a real visit"],
     "Career exploration across all domains", "Doctor, Engineer, Lawyer, Entrepreneur, Scientist, Designer, Manager",
     CLIENT + " (item 11) — matches old Q51 'career options'"),
]

LEARNING_STYLES_ORDER = ["Visual", "Aural", "Read/Write", "Kinesthetic", "Multimodal"]

LEARNING_DROPPED = [
    ("item 9 — 'At home, I am easily distracted by:'",
     "Not a VARK item. It is a distraction inventory, and option C (phone texts) has nothing to do with Read/Write."),
    ("items 1, 3, 4, 5, 6, 8", "Valid items, cut only because the blueprint caps learning styles at 4."),
]

# =========================================================================== #
# MULTIPLE INTELLLIGENCE — client's 4, used as-is. This sheet was already
# complete: 5 options, per-option intelligence, clusters and professions.
# =========================================================================== #
MI = [
    ("Q52", "Your teacher introduces a completely new topic in class. Which approach helps you understand it the fastest?",
     ["Understand the logic and steps behind it.", "Look at diagrams, charts, or visual examples.",
      "Listen carefully and discuss it with others.", "Think about it quietly and relate it to what I already know.",
      "Try a practical activity or experiment myself."],
     "STEM & Technology, Design, Education, Research, Healthcare", "Engineer, Architect, Teacher, Scientist, Surgeon"),
    ("Q53", "You have to remember an important lesson for next week's test. What do you naturally rely on?",
     ["Recognize patterns and understand how ideas connect.", "Picture the page, diagram, or image in my mind.",
      "Explain the lesson aloud or discuss it with someone.", "Connect the lesson to my own experiences or thoughts.",
      "Practise or apply what I learned."],
     "Engineering, Media, Psychology, Healthcare", "Data Scientist, Journalist, Psychologist, Physiotherapist"),
    ("Q54", "You receive a kit with no demonstration to learn a new activity. How do you begin?",
     ["Read the instructions and understand the sequence.", "Look at the illustrations before starting.",
      "Ask questions or discuss how it works.", "Think through the process before beginning.",
      "Start using it and learn through trial and practice."],
     "Engineering, Architecture, Technical Careers, Research", "Engineer, Architect, Product Designer, Technician"),
    ("Q55", "After learning something difficult, how do you know you've really understood it?",
     ["I can explain the reasoning behind it.", "I can visualize or sketch it clearly.",
      "I can explain it confidently to someone else.", "I can connect it with things I've learned before.",
      "I can use it correctly in a real task or activity."],
     "STEM, Design, Education, Healthcare", "Scientist, Graphic Designer, Teacher, Doctor"),
]

MI_ORDER = ["Logical-Mathematical", "Visual-Spatial", "Linguistic / Interpersonal",
            "Intrapersonal", "Bodily-Kinesthetic"]

# =========================================================================== #
# EMOTIONAL INTELLIGENCE — client's first 5, which map one-to-one onto the old
# Q56-Q60 dimensions. A-E map positionally to the 5 EQ domains (client design).
# =========================================================================== #
EI = [
    ("Q56", "Your teacher says your performance needs improvement. What would you most likely do?",
     ["Identify where I went wrong.", "Stay calm and avoid getting discouraged.",
      "Set a goal to improve next time.", "Understand the teacher's expectations.",
      "Ask for detailed feedback and guidance."],
     "Emotional Awareness", "MSCEIT + Goleman",
     "Education, Healthcare, Business, Leadership, Public Service", "Scientist, Teacher, Entrepreneur, Manager"),
    ("Q57", "Your preparation for an important competition is not going as planned. What would you do?",
     ["Recognize what is slowing me down.", "Stay calm and adjust my routine.",
      "Work harder with a fresh plan.", "Learn from teammates' experiences.",
      "Discuss strategies with mentors or friends."],
     "Emotional Regulation", "MSCEIT + Bar-On",
     "Sports, Business, Engineering, Research", "Doctor, Defence Officer, Engineer, Project Manager"),
    ("Q58", "A close friend performs poorly in an exam and feels upset. What would you do?",
     ["Think about how I would feel in that situation.", "Stay patient and avoid judging them.",
      "Encourage them to keep trying.", "Listen carefully before giving advice.",
      "Help them prepare for the next exam."],
     "Empathy & Social Awareness", "MSCEIT + Goleman",
     "Psychology, Education, Healthcare, HR", "Doctor, Teacher, Counsellor, Psychologist"),
    ("Q59", "Two classmates strongly disagree during a group project. What would you naturally do?",
     ["Reflect before sharing my opinion.", "Remain calm during the discussion.",
      "Focus on completing the task successfully.", "Listen to both viewpoints fairly.",
      "Help the team reach a solution together."],
     "Relationship Management", "Goleman + Genos",
     "Management, Law, Public Administration", "Manager, Lawyer, Consultant, IAS Officer"),
    ("Q60", "Your plans suddenly change because of an unexpected situation. How do you respond?",
     ["Understand how the change affects me.", "Accept the situation calmly.",
      "Look for a new opportunity.", "Consider how it affects others too.",
      "Work with others to adjust the plan."],
     "Adaptability & Resilience", "Bar-On + Genos",
     "Business, Defence, Healthcare, Technology", "Entrepreneur, Product Manager, Consultant, Business Leader"),
]

EI_ORDER = ["Self-Awareness", "Self-Regulation", "Self-Motivation", "Empathy", "Relationship Management"]

EI_GLOBAL_NOTE = (
    "SCORING RULE — this sheet yields an EQ PROFILE, not an EQ SCORE. Every option on every item is a healthy "
    "response (identify the error / stay calm / set a goal / understand them / ask for guidance), so there is no "
    "weak answer to score against and no defensible way to derive a high-or-low EQ level. What it does measure "
    "cleanly is which of the five domains the student reaches for first. Score it as forced choice: count how "
    "often each domain is chosen across the 5 items and report the leading one or two as a tendency. Do NOT "
    "publish a percentage or an out-of-100 EQ figure from these 5 items — with one observation per domain a "
    "single changed answer would move the result, so it is not reliable enough to quantify. This replaces the "
    "old set's graded 3/2/2/1 model, which the engine currently implements; the engine needs switching to "
    "domain counting. If you later want a genuine EQ level, that needs low-EQ distractors, which would mean "
    "changing the 60-question blueprint."
)

EI_DROPPED = [
    ("items 6-10 — praise for shared work; unexpected criticism; little time left; new student; after a hard task",
     "Valid items, cut only because the blueprint caps EI at 5. Client items 1-5 were kept because they map "
     "one-to-one onto the old Q56-Q60 dimensions."),
]

CLUSTER_NAMES = {
    "A": "Engineering & Construction", "B": "Information Technology", "C": "Health Science",
    "D": "Arts, Media & Design", "E": "Business & Marketing", "F": "Human & Public Services",
    "G": "Science, Nature & Agriculture", "H": "Sports, Hospitality & Lifestyle",
}

# --------------------------------------------------------------------------- #
# Styling helpers
# --------------------------------------------------------------------------- #
HEAD_FILL = PatternFill("solid", fgColor="1F3864")
HEAD_FONT = Font(bold=True, color="FFFFFF", size=10)
TOTAL_FILL = PatternFill("solid", fgColor="D9E2F3")
FLAG_FILL = PatternFill("solid", fgColor="FFF2CC")
NEW_FILL = PatternFill("solid", fgColor="E2EFDA")
REQ_FILL = PatternFill("solid", fgColor="FFC7CE")
BODY_FONT = Font(size=10)
THIN = Side(style="thin", color="BFBFBF")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TOP_WRAP = Alignment(vertical="top", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)


def per_option(labels):
    return "\n".join(f"{chr(65 + i)} · {v}" for i, v in enumerate(labels))


def points_str(d):
    """{'O': 3, 'C': 1} -> 'O+3, C+1'. Empty dict means the option earns nothing."""
    if not d:
        return "not scored"
    return ", ".join(f"{k}+{v}" for k, v in d.items())


def write_sheet(ws, headers, widths, rows, row_height=96):
    ws.append(headers)
    for c, (h, w) in enumerate(zip(headers, widths), 1):
        cell = ws.cell(row=1, column=c)
        cell.fill, cell.font, cell.alignment, cell.border = HEAD_FILL, HEAD_FONT, CENTER, BOX
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[1].height = 34
    for r in rows:
        ws.append(r)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(headers)):
        for cell in row:
            cell.font, cell.alignment, cell.border = BODY_FONT, TOP_WRAP, BOX
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = row_height
        ws.cell(row=r, column=1).alignment = CENTER
    ws.freeze_panes = "C2"


def shade_col(ws, col, predicate, fill):
    for r in range(2, ws.max_row + 1):
        cell = ws.cell(row=r, column=col)
        if predicate(cell.value):
            cell.fill = fill


# --------------------------------------------------------------------------- #
def build():
    wb = Workbook()
    wb.remove(wb.active)

    # ------------------------------- Blueprint ---------------------------- #
    ws = wb.create_sheet("Blueprint (60)")
    rows = [[rng.split("-")[0][1:], name, rng, count, 5, BLUEPRINT_NOTE[name]]
            for name, count, rng in BLUEPRINT]
    rows.append(["", "TOTAL", "Q1-Q60", sum(b[1] for b in BLUEPRINT), 5,
                 "8 sections · 60 questions · 5 options each — matches the old count exactly"])
    write_sheet(ws, ["#", "Sheet", "Q.No Range", "Questions", "Options / Q", "Construct Measured"],
                [5, 24, 13, 11, 12, 88], rows, row_height=30)
    for c in range(1, 7):
        cell = ws.cell(row=ws.max_row, column=c)
        cell.font, cell.fill = Font(bold=True, size=10), TOTAL_FILL

    # ------------------------------- interests ---------------------------- #
    ws = wb.create_sheet("interests")
    rows = []
    for qid, text, opts, source, note in INTERESTS:
        rows.append([
            qid, text, *[o[0] for o in opts],
            per_option([f"{o[1]} — {CLUSTER_NAMES[o[1]]}" for o in opts]),
            per_option([o[2] for o in opts]),
            # dominant RIASEC letter first
            per_option(["+".join(sorted(o[3], key=o[3].get, reverse=True)) for o in opts]),
            per_option([points_str(o[3]) for o in opts]),
            source, note,
        ])
    write_sheet(ws, [
        "Q.No", "Situation", "Option A", "Option B", "Option C", "Option D", "Option E",
        "Career Cluster (per option)", "Primary Career Matches (per option)",
        "RIASEC Code (per option)", "RIASEC Weights (per option)", "Source", "Review Note",
    ], [7, 40, 30, 30, 30, 30, 30, 30, 50, 16, 24, 26, 46], rows, row_height=112)
    shade_col(ws, 12, lambda v: v.startswith(AUTHORED) if isinstance(v, str) else False, NEW_FILL)
    shade_col(ws, 13, lambda v: bool(v), FLAG_FILL)
    r = ws.max_row + 2
    ws.cell(row=r, column=1, value="NOTE").font = Font(bold=True, size=10)
    n = ws.cell(row=r, column=2, value=INTERESTS_SCORING_NOTE)
    n.alignment, n.font, n.fill = TOP_WRAP, BODY_FONT, FLAG_FILL
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=13)
    ws.row_dimensions[r].height = 60

    # ------------------------------- aptitude ----------------------------- #
    ws = wb.create_sheet("aptitude")
    rows = []
    for (qid, text, opts, correct, dim, clusters, profs, img, spec, source, note) in APTITUDE:
        rows.append([qid, text, *opts, correct, dim, img, spec, clusters, profs, source, note])
    write_sheet(ws, [
        "Q.No", "Question", "Option A", "Option B", "Option C", "Option D", "Option E",
        "Correct Answer", "Aptitude Dimension", "Image", "Image / Visual Specification",
        "Career Clusters", "Representative Professions", "Source", "Review Note",
    ], [7, 44, 16, 16, 18, 16, 18, 9, 20, 12, 60, 26, 30, 22, 60], rows, row_height=104)
    for r in range(2, ws.max_row + 1):
        c8 = ws.cell(row=r, column=8)
        c8.alignment, c8.font = CENTER, Font(bold=True, size=10)
        c10 = ws.cell(row=r, column=10)
        c10.alignment = CENTER
        if c10.value == "Required":
            c10.font, c10.fill = Font(bold=True, size=10, color="9C0006"), REQ_FILL
        elif c10.value == "Recommended":
            c10.fill = FLAG_FILL
    shade_col(ws, 15, lambda v: bool(v), FLAG_FILL)

    # ------------------------------- personality -------------------------- #
    ws = wb.create_sheet("personality")
    rows = []
    for qid, text, opts, trait, facet, note in PERSONALITY:
        rows.append([
            qid, text, *[t for t, _p in opts],
            trait, facet,
            per_option([points_str(p) for _t, p in opts]),
            CLIENT + " (options expanded)", note,
        ])
    write_sheet(ws, [
        "Q.No", "Situation", "Option A", "Option B", "Option C", "Option D", "Option E",
        "Big Five Trait", "Personality Facet", "Trait Weights (per option)", "Source", "Review Note",
    ], [7, 44, 30, 30, 30, 30, 16, 18, 18, 26, 24, 56], rows, row_height=100)
    shade_col(ws, 12, lambda v: bool(v), FLAG_FILL)
    r = ws.max_row + 2
    ws.cell(row=r, column=1, value="NOTE").font = Font(bold=True, size=10)
    n = ws.cell(row=r, column=2, value=PERSONALITY_GLOBAL_NOTES)
    n.alignment, n.font, n.fill = TOP_WRAP, BODY_FONT, FLAG_FILL
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=12)
    ws.row_dimensions[r].height = 56

    # ------------------------------- Strenghts ---------------------------- #
    ws = wb.create_sheet("Strenghts")
    rows = []
    for qid, text, opts in STRENGTHS:
        domains = [max(p, key=p.get) for _t, p in opts]
        clusters, profs = STRENGTHS_CAREERS[qid]
        rows.append([
            qid, text, *[t for t, _p in opts],
            ", ".join(dict.fromkeys(domains)),
            per_option(domains),
            per_option([points_str(p) for _t, p in opts]),
            clusters, profs, AUTHORED,
        ])
    write_sheet(ws, [
        "Q.No", "Question", "Option A", "Option B", "Option C", "Option D", "Option E",
        "Strength Domains Measured", "Strength Domain (per option)", "Strength Weights (per option)",
        "Career Clusters", "Example Professions", "Source",
    ], [7, 42, 32, 32, 32, 32, 32, 30, 22, 32, 32, 38, 22], rows, row_height=110)
    shade_col(ws, 13, lambda v: v == AUTHORED, NEW_FILL)

    # ------------------------------- motivators --------------------------- #
    ws = wb.create_sheet("motivators")
    rows = []
    for qid, text, opts in MOTIVATORS:
        domains = [max(p, key=p.get) for _t, p, _s in opts]
        clusters, profs = MOTIVATOR_CAREERS[qid]
        rows.append([
            qid, text, *[t for t, _p, _s in opts],
            ", ".join(dict.fromkeys(domains)),
            per_option(domains),
            per_option([points_str(p) for _t, p, _s in opts]),
            clusters, profs,
            per_option([s for _t, _p, s in opts]),
        ])
    write_sheet(ws, [
        "Q.No", "Question", "Option A", "Option B", "Option C", "Option D", "Option E",
        "Motivator Dimensions Measured", "Motivator (per option)", "Motivator Weights (per option)",
        "Career Clusters", "Example Professions", "Source (per option)",
    ], [7, 44, 30, 30, 30, 30, 32, 30, 20, 34, 36, 38, 26], rows, row_height=110)

    # ------------------------------- Learning styles ---------------------- #
    ws = wb.create_sheet("Learning styles")
    rows = []
    for qid, text, opts, clusters, profs, source in LEARNING:
        rows.append([qid, text, *opts, per_option(LEARNING_STYLES_ORDER),
                     " / ".join(LEARNING_STYLES_ORDER), clusters, profs, source])
    write_sheet(ws, [
        "Q.No", "Scenario",
        "Option A (Visual)", "Option B (Aural)", "Option C (Read/Write)",
        "Option D (Kinesthetic)", "Option E (Multimodal)",
        "Learning Preference (per option)", "Preferences Measured",
        "Career Clusters", "Example Professions", "Source",
    ], [7, 44, 28, 28, 30, 30, 38, 22, 40, 38, 44, 48], rows, row_height=100)

    # ------------------------------- multiple intellligence --------------- #
    ws = wb.create_sheet("multiple intellligence")
    rows = []
    for qid, text, opts, clusters, profs in MI:
        rows.append([qid, text, *opts, per_option(MI_ORDER), ", ".join(MI_ORDER),
                     clusters, profs, CLIENT])
    write_sheet(ws, [
        "Q.No", "Question", "Option A", "Option B", "Option C", "Option D", "Option E",
        "Primary Intelligence (per option)", "Intelligences Measured",
        "Career Clusters", "Example Professions", "Source",
    ], [7, 44, 32, 32, 32, 34, 32, 26, 44, 34, 40, 20], rows, row_height=104)

    # ------------------------------- emotional intelligence --------------- #
    ws = wb.create_sheet("emotional intelligence")
    rows = []
    for qid, text, opts, dim, framework, clusters, profs in EI:
        rows.append([qid, text, *opts, per_option(EI_ORDER), dim, framework, clusters, profs, CLIENT])
    write_sheet(ws, [
        "Q.No", "Question",
        "Option A (Self-Awareness)", "Option B (Self-Regulation)", "Option C (Self-Motivation)",
        "Option D (Empathy)", "Option E (Relationship Mgmt)",
        "EQ Domain (per option)", "Question Dimension (old set)", "Framework Inspiration",
        "Career Clusters", "Example Professions", "Source",
    ], [7, 42, 28, 28, 28, 28, 30, 24, 22, 20, 30, 36, 20], rows, row_height=100)
    r = ws.max_row + 2
    ws.cell(row=r, column=1, value="NOTE").font = Font(bold=True, size=10)
    n = ws.cell(row=r, column=2, value=EI_GLOBAL_NOTE)
    n.alignment, n.font, n.fill = TOP_WRAP, BODY_FONT, FLAG_FILL
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=13)
    ws.row_dimensions[r].height = 72

    # ------------------------------- Review Notes ------------------------- #
    ws = wb.create_sheet("Review Notes")
    notes = []
    notes.append(["Aptitude", "Dropped items (blueprint caps aptitude at 10)", "", ""])
    for what, why in APTITUDE_DROPPED:
        notes.append(["", what, why, ""])
    notes.append(["Learning styles", "Dropped items (blueprint caps at 4)", "", ""])
    for what, why in LEARNING_DROPPED:
        notes.append(["", what, why, ""])
    notes.append(["Emotional intelligence", "Dropped items (blueprint caps at 5)", "", ""])
    for what, why in EI_DROPPED:
        notes.append(["", what, why, ""])
    notes.append(["Emotional intelligence", "Scoring model — resolved", EI_GLOBAL_NOTE, "Decided — review"])
    notes.append(["Personality", "Unscored option E — resolved", PERSONALITY_GLOBAL_NOTES, "Decided — review"])
    notes.append(["Personality", "Option wording", "The client's options were shorthand notes ('Clearest, idea most "
                                                   "reliable approach'), not usable by a Class 9-10 student. All 60 "
                                                   "options expanded into full sentences with the meaning unchanged. "
                                                   "Original wording preserved in the client-original workbook.",
                  "Done — review"])
    notes.append(["Personality", "Trait weights", "Derived for all 12 items from the client's option meanings, using "
                                                  "the old set's convention (primary trait +3, supporting +1/+2). "
                                                  "Trait and facet carried over from the old set, except Q28 "
                                                  "(Confidence to Initiative) and Q30/Q31 (Empathy and Cooperation "
                                                  "swapped to match the client's scenarios).", "Done — review"])
    notes.append(["Interests", "RIASEC codes", "Derived for all 12 items from what each option actually involves "
                                              "(the client sheet gave career matches but no RIASEC letters). "
                                              "Dominant code listed first.", "Done — review"])
    notes.append(["Interests", "Scoring rule", INTERESTS_SCORING_NOTE, "Decided — review"])
    notes.append(["Interests", "Q7 option C changed", "Client option C (legal protocols / Law & Administration) "
                                                      "duplicated option B's cluster, so two of five choices gave the "
                                                      "same signal and Science/Agriculture was unmeasured. Replaced "
                                                      "with a lab/epidemiology option.", "Changed — review"])
    notes.append(["Interests", "Q9 option D narrowed", "Client option D mixed Law and Creative Arts in one choice, "
                                                       "making the answer ambiguous between two clusters. Narrowed to "
                                                       "media/publishing; Law is covered on six other items.",
                  "Changed — review"])
    notes.append(["Aptitude", "Q19 cipher fixed", "Client keys contradicted each other (TECH gave T=3, ARTS gave "
                                                  "T=7) and the marked answer assumed T=7. Second key changed to "
                                                  "ARTS=5637 so T=3 throughout; RATE = 6531.", "Fixed"])
    notes.append(["Aptitude", "Q20 series fixed", "Client series started at circles, so the count rule (+2) and the "
                                                 "side rule (+1) disagreed and option D was arguable. Series now "
                                                 "starts at a triangle; answer is 9 hexagons.", "Fixed"])
    notes.append(["Aptitude", "Cluster tags retagged", "Six items carried clusters read off the correct answer rather "
                                                       "than the skill (coding-decoding tagged 'Genetic Coder', "
                                                       "paper-folding tagged 'Orthopedic Specialist', odd-one-out "
                                                       "tagged 'Farm Operator'). Retagged from the old bank's "
                                                       "dimension-to-cluster table.", "Fixed"])
    notes.append(["Interests", "Source format", "The client interests sheet was a raw CSV paste — question number and "
                                               "text in one cell, options as 'A,text' in a single column, unbalanced "
                                               "quote marks and a stray leading comma on Q1 option C. Restructured "
                                               "into proper columns here.", "Fixed"])
    notes.append(["Strengths", "Construct", "The client sheet measured career interests, not strengths: all 12 items "
                                           "reused the same Tech / Health / Agri-Eco / Law-Finance buckets as the "
                                           "interests sheet, and every option was prefixed with its own bucket name, "
                                           "which telegraphs the answer. Rebuilt with 8 items covering the 8 strength "
                                           "domains.", "Rebuilt — review"])
    notes.append(["Aptitude", "Junk rows", "Client sheet row 61 contained a stray 'E | E' and row 92 was blank; both "
                                          "dropped. Items were also numbered 11-20 then restarting at 11-15, so "
                                          "11-15 appeared twice. Renumbered Q13-Q22.", "Fixed"])
    notes.append(["Aptitude", "Correct-answer marks", "Client sheet used two different marks (checkbox in the first "
                                                     "block, green tick in the second). Replaced with a single "
                                                     "Correct Answer column.", "Fixed"])
    notes.append(["Aptitude", "Images", "3 items need artwork before launch (Q17 paper-fold, Q20 shape series, "
                                       "Q22 gears) and 1 is recommended (Q16 kWh bar chart). Q22's artwork already "
                                       "exists as inline SVG in the old bank and can be reused.", "Action needed"])
    notes.append(["Engine", "Option count", "The app scoring code types these items as 'choice4' and reads exactly "
                                           "four options. It must be widened to five before this set can be "
                                           "imported.", "Action needed"])
    write_sheet(ws, ["Sheet", "Item", "Detail", "Status"], [22, 54, 96, 20], notes, row_height=46)
    shade_col(ws, 4, lambda v: v in ("Needs a decision", "Needs input", "Action needed"), FLAG_FILL)
    shade_col(ws, 4, lambda v: v == "Fixed", NEW_FILL)

    # ------------------------------- verify ------------------------------- #
    problems = []
    grand = 0
    for name, count, _rng in BLUEPRINT:
        sheet = wb[name]
        data = [r for r in sheet.iter_rows(min_row=2, values_only=True)
                if r[0] and str(r[0]).startswith("Q")]
        grand += len(data)
        if len(data) != count:
            problems.append(f"{name}: {len(data)} questions, blueprint says {count}")
        for row in data:
            if any(row[c] in (None, "") for c in range(2, 7)):
                problems.append(f"{name} {row[0]}: fewer than 5 options")
    if grand != 60:
        problems.append(f"total is {grand}, expected 60")
    if problems:
        raise SystemExit("VERIFY FAILED:\n  " + "\n  ".join(problems))

    TARGET.parent.mkdir(parents=True, exist_ok=True)
    wb.save(TARGET)
    print(f"wrote {TARGET.relative_to(PROJECT)}")
    for name, count, rng in BLUEPRINT:
        print(f"  {name:<24} {count:>2} questions  {rng}")
    print(f"  {'TOTAL':<24} {grand:>2} questions  · 5 options each")


# --------------------------------------------------------------------------- #
# Markdown export. GitHub cannot render .xlsx, so the same 60 questions are also
# written to a reviewable Markdown file next to the workbook.
# --------------------------------------------------------------------------- #
def write_markdown(path: Path):
    L: list[str] = []
    L.append("# Finalised Question Set 2026")
    L.append("")
    L.append("60 questions · 5 options each · classes 9-10. Generated by "
             "`project/scripts/build_finalised_set_2026.py` — edit the script, not this file.")
    L.append("")
    L.append("| # | Section | Q.No | Questions | Options | Construct |")
    L.append("|---|---|---|---|---|---|")
    for i, (name, count, rng) in enumerate(BLUEPRINT, 1):
        L.append(f"| {i} | {name} | {rng} | {count} | 5 | {BLUEPRINT_NOTE[name]} |")
    L.append(f"| | **TOTAL** | Q1-Q60 | **{sum(b[1] for b in BLUEPRINT)}** | 5 | |")
    L.append("")

    def opts_block(texts, tags=None):
        for i, t in enumerate(texts):
            tag = f" — _{tags[i]}_" if tags else ""
            L.append(f"- **{chr(65 + i)}.** {t}{tag}")
        L.append("")

    # -------- interests
    L.append("---")
    L.append("")
    L.append("## 1. Interests — Q1-Q12")
    L.append("")
    L.append(f"> {INTERESTS_SCORING_NOTE}")
    L.append("")
    for qid, text, opts, source, note in INTERESTS:
        L.append(f"### {qid}. {text}")
        L.append("")
        opts_block([o[0] for o in opts],
                   [f"{o[1]} · {CLUSTER_NAMES[o[1]]} · RIASEC {points_str(o[3])}" for o in opts])
        L.append(f"_Source: {source}_")
        if note:
            L.append("")
            L.append(f"> **Note.** {note}")
        L.append("")

    # -------- aptitude
    L.append("---")
    L.append("")
    L.append("## 2. Aptitude — Q13-Q22")
    L.append("")
    L.append("Artwork status is in the **Image** line. `Required` means the item cannot be answered without it.")
    L.append("")
    for (qid, text, opts, correct, dim, clusters, profs, img, spec, source, note) in APTITUDE:
        L.append(f"### {qid}. {text}")
        L.append("")
        opts_block([f"{o}{'  ✅' if chr(65 + n) == correct else ''}" for n, o in enumerate(opts)])
        L.append(f"- **Answer:** {correct}")
        L.append(f"- **Dimension:** {dim}")
        L.append(f"- **Image:** {img} — {spec}")
        L.append(f"- **Clusters:** {clusters}")
        L.append(f"- **Professions:** {profs}")
        L.append(f"- _Source: {source}_")
        if note:
            L.append("")
            L.append(f"> **Note.** {note}")
        L.append("")

    # -------- personality
    L.append("---")
    L.append("")
    L.append("## 3. Personality — Q23-Q34")
    L.append("")
    L.append(f"> {PERSONALITY_GLOBAL_NOTES}")
    L.append("")
    for qid, text, opts, trait, facet, note in PERSONALITY:
        L.append(f"### {qid}. {text}")
        L.append("")
        opts_block([t for t, _p in opts], [points_str(p) for _t, p in opts])
        L.append(f"- **Trait / facet:** {trait} · {facet}")
        if note:
            L.append("")
            L.append(f"> **Note.** {note}")
        L.append("")

    # -------- strengths
    L.append("---")
    L.append("")
    L.append("## 4. Strengths — Q35-Q42")
    L.append("")
    L.append("> Rebuilt to measure the 8 strength domains. The client's own sheet reused the "
             "interests buckets (Tech / Health / Agri / Law) and labelled every option with its "
             "bucket, so it measured career interest, not strengths.")
    L.append("")
    for qid, text, opts in STRENGTHS:
        clusters, profs = STRENGTHS_CAREERS[qid]
        L.append(f"### {qid}. {text}")
        L.append("")
        opts_block([t for t, _p in opts], [points_str(p) for _t, p in opts])
        L.append(f"- **Clusters:** {clusters}")
        L.append(f"- **Professions:** {profs}")
        L.append("- _Source: authored — please review_")
        L.append("")

    # -------- motivators
    L.append("---")
    L.append("")
    L.append("## 5. Motivators — Q43-Q47")
    L.append("")
    for qid, text, opts in MOTIVATORS:
        clusters, profs = MOTIVATOR_CAREERS[qid]
        L.append(f"### {qid}. {text}")
        L.append("")
        opts_block([t for t, _p, _s in opts],
                   [f"{points_str(p)} · {s}" for _t, p, s in opts])
        L.append(f"- **Clusters:** {clusters}")
        L.append(f"- **Professions:** {profs}")
        L.append("")

    # -------- learning styles
    L.append("---")
    L.append("")
    L.append("## 6. Learning Styles — Q48-Q51")
    L.append("")
    for qid, text, opts, clusters, profs, source in LEARNING:
        L.append(f"### {qid}. {text}")
        L.append("")
        opts_block(opts, LEARNING_STYLES_ORDER)
        L.append(f"- **Clusters:** {clusters}")
        L.append(f"- _Source: {source}_")
        L.append("")

    # -------- multiple intelligence
    L.append("---")
    L.append("")
    L.append("## 7. Multiple Intelligence — Q52-Q55")
    L.append("")
    L.append("> The one section that arrived complete — used exactly as supplied.")
    L.append("")
    for qid, text, opts, clusters, profs in MI:
        L.append(f"### {qid}. {text}")
        L.append("")
        opts_block(opts, MI_ORDER)
        L.append(f"- **Clusters:** {clusters}")
        L.append(f"- **Professions:** {profs}")
        L.append("")

    # -------- emotional intelligence
    L.append("---")
    L.append("")
    L.append("## 8. Emotional Intelligence — Q56-Q60")
    L.append("")
    L.append(f"> {EI_GLOBAL_NOTE}")
    L.append("")
    for qid, text, opts, dim, framework, clusters, profs in EI:
        L.append(f"### {qid}. {text}")
        L.append("")
        opts_block(opts, EI_ORDER)
        L.append(f"- **Question dimension (old set):** {dim} · _{framework}_")
        L.append(f"- **Clusters:** {clusters}")
        L.append("")

    # -------- open items
    L.append("---")
    L.append("")
    L.append("## Open items before launch")
    L.append("")
    L.append("| Area | Item | Status |")
    L.append("|---|---|---|")
    for qid, _t, _o, _c, _d, _cl, _p, img, _s, _src, _n in APTITUDE:
        if img in ("Required", "Recommended"):
            L.append(f"| Aptitude | {qid} artwork | {img} |")
    L.append("| Bank import | `abstainIndex: 4` on the 12 personality items | Not yet written |")
    L.append("| Bank import | `optionDomains` on the 5 EI items | Not yet written |")
    L.append("")
    L.append("Dropped items and every correction are listed on the **Review Notes** tab of the workbook.")
    L.append("")

    path.write_text("\n".join(L), encoding="utf-8")
    return path


if __name__ == "__main__":
    if len(sys.argv) > 1:
        TARGET = Path(sys.argv[1]).resolve()
        MIRROR = None
    build()
    md = write_markdown(TARGET.with_suffix(".md"))
    print(f"wrote {md.relative_to(PROJECT) if md.is_relative_to(PROJECT) else md}")
    if MIRROR:
        try:
            shutil.copy2(TARGET, MIRROR)
            print(f"mirrored to {MIRROR}")
        except PermissionError:
            print(f"could not refresh {MIRROR.name} — it is open in Excel (repo copy is current)")
