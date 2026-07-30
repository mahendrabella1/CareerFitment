"""Export the full Class 9-10 question paper to a single .xlsx workbook.

Reads the same three JSON banks the exam itself serves (lib/newAssessment/data.ts
loads these), filters to the "9-10" life stage, and writes one sheet per
category plus an Overview sheet.

Every set is included, not just the one a student happens to be served -- the
exam picks one set per category at random, so the paper is the whole pool.

    python scripts/export_9_10_question_paper.py

Output: ../Class_9_10_Question_Paper.xlsx (repo root, beside the source .xlsx
files). Re-run after editing any question bank to refresh it.
"""

import json
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

STAGE = "9-10"
HERE = Path(__file__).resolve().parent
DATA = HERE.parent / "data"
OUT = HERE.parent.parent / "Class_9_10_Question_Paper.xlsx"

LETTERS = "ABCDE"

# Section order matches CATEGORY_ORDER in lib/newAssessment/data.ts -- the order
# students actually sit them in.
CATEGORIES = [
    ("personality", "Personality", "assessment-questions.json"),
    ("career_interest", "Career Interests", "assessment-questions.json"),
    ("multiple_intelligence", "Multiple Intelligences", "assessment-questions.json"),
    ("emotional_intelligence", "Emotional Intelligence", "assessment-questions.json"),
    ("learning_styles", "Learning Style", "assessment-questions.json"),
    ("motivators", "Motivators", "assessment-questions.json"),
    ("strengths", "Strengths", "strengths-questions.json"),
    ("aptitude", "Aptitude", "aptitude-questions.json"),
]

# Cluster codes carried on career_interest options, spelled out so the sheet is
# readable without cross-referencing career-clusters.json.
CLUSTER_NAMES = {}


def load_banks():
    with open(DATA / "assessment-questions.json", encoding="utf-8") as f:
        bank = json.load(f)
    with open(DATA / "aptitude-questions.json", encoding="utf-8") as f:
        bank["aptitude"] = json.load(f)
    with open(DATA / "strengths-questions.json", encoding="utf-8") as f:
        bank["strengths"] = json.load(f)
    try:
        with open(DATA / "career-clusters.json", encoding="utf-8") as f:
            for code, v in json.load(f).items():
                CLUSTER_NAMES[code] = v.get("cluster", code)
    except FileNotFoundError:
        pass
    return bank


def natural_set_key(name):
    """'Set 10' must sort after 'Set 2'."""
    m = re.search(r"(\d+)", name)
    return (int(m.group(1)) if m else 0, name)


def html_to_text(html):
    """Flatten a media stimulus block into one readable cell."""
    s = re.sub(r"</li\s*>", " | ", html, flags=re.I)
    s = re.sub(r"</(p|div|ul|ol|tr)\s*>", " — ", s, flags=re.I)
    s = re.sub(r"<[^>]+>", " ", s)
    s = (s.replace("&nbsp;", " ").replace("&amp;", "&")
          .replace("&lt;", "<").replace("&gt;", ">").replace("&quot;", '"'))
    s = re.sub(r"\s+", " ", s).strip()
    return re.sub(r"(\s*[—|]\s*)+$", "", s).strip()


def stimulus_of(q):
    """Question setup shown above the options: a passage, or a visual grid."""
    media = q.get("media")
    if not media:
        return ""
    if isinstance(media, str):
        try:
            media = json.loads(media)
        except (ValueError, TypeError):
            return html_to_text(media)
    if not isinstance(media, dict):
        return ""
    kind = media.get("type")
    if kind == "html":
        return html_to_text(media.get("html", ""))
    if kind == "grid":
        n = len(media.get("cells") or [])
        return f"[Visual: pattern grid of {n} cells — see the live exam]"
    return f"[Visual: {kind}]"


def option_text(q, i):
    opts = q.get("options") or []
    if i >= len(opts):
        return ""
    val = opts[i]
    if not isinstance(val, str):
        return str(val)
    if q.get("svgOptions") or val.lstrip().startswith("<svg"):
        return f"[Shape {LETTERS[i]} — diagram]"
    return val


def mapping(q, key, pretty=None):
    """Render per-option scoring/tagging as 'A=x, B=y' in a single cell."""
    vals = q.get(key)
    if not isinstance(vals, list) or not vals:
        return ""
    out = []
    for i, v in enumerate(vals[: len(LETTERS)]):
        # A blank code means the option scores nothing (e.g. the "none of these
        # feel like me" escape hatch on most/least items).
        label = (pretty.get(v, v) if pretty else v) if v != "" else "(scores nothing)"
        out.append(f"{LETTERS[i]}={label}")
    return ", ".join(out)


TYPE_LABELS = {
    "choice4": "Single choice",
    "choice5": "Single choice",
    "vark": "Single choice",
    "mcq": "MCQ — one correct",
    "mostleast": "Most / Least like me",
    "open": "Open response (free text)",
}


def type_label(q):
    t = q.get("type", "")
    return TYPE_LABELS.get(t, t)


def correct_letter(q):
    c = q.get("correct")
    if not isinstance(c, int) or c < 0 or c >= len(LETTERS):
        return ""
    return f"{LETTERS[c]}. {option_text(q, c)}"


# Per-category trailing columns: (header, value function).
EXTRAS = {
    "personality": [
        ("Trait", lambda q: q.get("trait", "")),
        ("Reverse scored", lambda q: "Yes" if q.get("reverse") else "No"),
        ("Points per option", lambda q: mapping(q, "points")),
    ],
    "career_interest": [
        ("Cluster per option", lambda q: mapping(q, "clusters", CLUSTER_NAMES)),
    ],
    "multiple_intelligence": [
        ("Intelligence", lambda q: q.get("intelligence", "")),
    ],
    "emotional_intelligence": [
        ("EI domain", lambda q: q.get("domain", "")),
        ("Points per option", lambda q: mapping(q, "scores")),
    ],
    "learning_styles": [
        ("Style per option", lambda q: mapping(q, "styles")),
    ],
    "motivators": [
        ("Motivator per option", lambda q: mapping(q, "domains")),
    ],
    "strengths": [
        # "strengths_selfreport" is an internal bucket, not a readable domain.
        ("Domain", lambda q: "Self-report" if q.get("domain") == "strengths_selfreport"
                             else q.get("domain", "")),
        # Only the "most/least like me" items tag options with a strength code.
        ("Strength per option", lambda q: mapping(q, "domains")),
        ("Correct answer", correct_letter),
    ],
    "aptitude": [
        ("Domain", lambda q: q.get("domain", "")),
        ("Format", lambda q: q.get("format", "")),
        ("Correct answer", correct_letter),
    ],
}

HEAD_FILL = PatternFill("solid", fgColor="1E1B4B")
HEAD_FONT = Font(bold=True, color="FFFFFF", size=10.5)
SET_FILL = PatternFill("solid", fgColor="EEF0FF")
THIN = Side(style="thin", color="D8DAE5")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
TOP_WRAP = Alignment(vertical="top", wrap_text=True)
TOP_CENTER = Alignment(vertical="top", horizontal="center")


def write_sheet(wb, key, title, sets):
    ws = wb.create_sheet(title[:31])

    max_opts = max((len(q.get("options") or []) for qs in sets.values() for q in qs), default=4)
    has_stimulus = any(stimulus_of(q) for qs in sets.values() for q in qs)

    headers = ["Set", "Q#", "Type"]
    if has_stimulus:
        headers.append("Setup / stimulus")
    headers.append("Question")
    headers += [f"Option {LETTERS[i]}" for i in range(max_opts)]
    headers += [h for h, _ in EXTRAS.get(key, [])]

    ws.append(headers)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill, cell.font = HEAD_FILL, HEAD_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30

    total = 0
    for set_name in sorted(sets, key=natural_set_key):
        questions = sets[set_name]
        for n, q in enumerate(questions, start=1):
            row = [set_name, q.get("q") or f"Q{n}", type_label(q)]
            if has_stimulus:
                row.append(stimulus_of(q))
            row.append(q.get("text", ""))
            row += [option_text(q, i) for i in range(max_opts)]
            row += [fn(q) for _, fn in EXTRAS.get(key, [])]
            ws.append(row)
            total += 1

            r = ws.max_row
            for c in range(1, len(headers) + 1):
                cell = ws.cell(row=r, column=c)
                cell.alignment = TOP_WRAP
                cell.border = BORDER
            ws.cell(row=r, column=2).alignment = TOP_CENTER
            # Tint the first row of each set so the blocks are scannable.
            if n == 1:
                for c in range(1, len(headers) + 1):
                    ws.cell(row=r, column=c).fill = SET_FILL

    widths = {"Set": 9, "Q#": 6, "Type": 20, "Setup / stimulus": 52, "Question": 60}
    for i, h in enumerate(headers, start=1):
        if h.startswith("Option"):
            w = 34
        elif h in ("Points per option", "Cluster per option", "Style per option",
                   "Motivator per option", "Strength per option", "Correct answer"):
            w = 30
        else:
            w = widths.get(h, 20)
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{ws.max_row}"
    return total, len(sets)


def write_overview(wb, rows, grand_total):
    ws = wb.create_sheet("Overview", 0)
    ws["A1"] = "OneGrasp — Class 9–10 Question Paper"
    ws["A1"].font = Font(bold=True, size=15, color="1E1B4B")
    ws["A2"] = ("Every question in the Class 9–10 bank, all sets. The live exam serves "
                "ONE randomly picked set per section, so a student sees roughly an eighth of this.")
    ws["A2"].font = Font(size=10, color="5B6070")
    ws.merge_cells("A2:E2")
    ws.row_dimensions[2].height = 28
    ws["A2"].alignment = Alignment(vertical="center", wrap_text=True)

    head = ["#", "Section", "Sets", "Questions in bank", "Per student (1 set)"]
    ws.append([])
    ws.append(head)
    hr = ws.max_row
    for c in range(1, len(head) + 1):
        cell = ws.cell(row=hr, column=c)
        cell.fill, cell.font = HEAD_FILL, HEAD_FONT
        cell.alignment = Alignment(vertical="center")

    for i, (title, n_sets, n_q) in enumerate(rows, start=1):
        per = round(n_q / n_sets) if n_sets else 0
        ws.append([i, title, n_sets, n_q, per])
        for c in range(1, len(head) + 1):
            ws.cell(row=ws.max_row, column=c).border = BORDER

    ws.append(["", "TOTAL", "", grand_total, sum(round(q / s) if s else 0 for _, s, q in rows)])
    for c in range(1, len(head) + 1):
        cell = ws.cell(row=ws.max_row, column=c)
        cell.font = Font(bold=True)
        cell.fill = SET_FILL

    note_row = ws.max_row + 2
    ws.cell(row=note_row, column=1, value="Notes").font = Font(bold=True, size=11, color="1E1B4B")
    notes = [
        "Aptitude and Strengths have objectively correct answers — see the "
        "'Correct answer' column. Every other section is self-report and has no right answer.",
        "Visual questions (pattern matrices, shape series) can't be shown as text. Their options "
        "read '[Shape A — diagram]' and their stimulus reads '[Visual: pattern grid …]'.",
        "The 10 'Open response' items in Strengths (one per set) have no options by design — the "
        "student types a free-text answer. Their Option columns are intentionally blank.",
        "'Points per option' is the score each choice contributes. 'Cluster / Style / Motivator "
        "per option' shows which trait an option feeds, not a score.",
        "Reverse-scored personality items are flagged; their points are already stored in the "
        "final direction, so no extra inversion is applied at scoring time.",
        "Regenerate with: python scripts/export_9_10_question_paper.py",
    ]
    for i, n in enumerate(notes):
        r = note_row + 1 + i
        ws.cell(row=r, column=1, value="•")
        c = ws.cell(row=r, column=2, value=n)
        c.alignment = Alignment(vertical="top", wrap_text=True)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
        ws.row_dimensions[r].height = 30

    for col, w in zip("ABCDE", (6, 34, 10, 20, 20)):
        ws.column_dimensions[col].width = w


def main():
    bank = load_banks()
    wb = Workbook()
    wb.remove(wb.active)

    summary, grand_total = [], 0
    for key, title, _src in CATEGORIES:
        sets = (bank.get(key) or {}).get(STAGE) or {}
        if not sets:
            print(f"  ! no '{STAGE}' data for {key} — skipped")
            continue
        n_q, n_sets = write_sheet(wb, key, title, sets)
        summary.append((title, n_sets, n_q))
        grand_total += n_q
        print(f"  {title:<26} {n_sets:>2} sets   {n_q:>4} questions")

    write_overview(wb, summary, grand_total)
    wb.save(OUT)
    print(f"\n{grand_total} questions across {len(summary)} sections -> {OUT}")


if __name__ == "__main__":
    main()
